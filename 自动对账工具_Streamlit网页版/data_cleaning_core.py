from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


STANDARD_COLUMNS = [
    "暂留备用",
    "批准文号",
    "药品名称",
    "最终药品编码",
    "批号",
    "生产日期",
    "有效期至",
    "数量",
    "上调后单价",
    "供货商名称",
    "供货商名称编码",
    "供货单位名称",
    "供货单位名称编码",
    "采购人编码",
    "生产厂家",
    "单价",
    "规格",
    "厂家批准文号清洗",
    "厂家商品规格清洗",
    "单位",
    "厂家包装单位清洗",
    "单价上调类型",
    "单价上调值/比例",
    "单价上调表中的药品编码",
    "候选规格",
    "候选单位",
]

DEFAULT_ALIASES = {
    "日期": ["日期", "销售日期", "单据日期", "出库日期"],
    "供货单位名称": ["单位名称", "单位全名", "客户名称", "部门名称"],
    "药品名称": ["药品名称", "通用名称", "商品名称", "通用名/商品名", "通用名称", "品名", "名称"],
    "规格": ["规格", "商品规格", "药品规格", "规格/型号"],
    "批号": ["批号", "产品批号", "生产批号", "批号/序列号"],
    "数量": ["数量", "销售数量"],
    "单价": ["单价", "含税价", "含税单价", "供价", "价格"],
    "金额": ["金额", "含税金额"],
    "生产厂家": ["生产厂家", "产地/生产厂家", "生产企业", "生产厂商", "厂商全名", "产地"],
    "批准文号": ["批准文号"],
    "单位": ["单位", "包装单位", "基本单位"],
    "生产日期": ["生产日期"],
    "有效期至": ["有效期至", "有效期"],
}


@dataclass
class CleanResult:
    file_name: str
    supplier: str
    total_rows: int
    success_rows: int
    failed_rows: int


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()


def normalize_key(value: object) -> str:
    return re.sub(r"[\s\-_()（）/\\*×xX,，.。:：;；]+", "", normalize_text(value)).lower()


def normalize_approval(value: object) -> str:
    text = normalize_text(value)
    text = text.replace("准字", "准字").replace("国药准字 ", "国药准字")
    return re.sub(r"[\s\-()（）]+", "", text).upper()


def to_decimal(value: object) -> Optional[Decimal]:
    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError):
        return None


def read_excel_any(path: Path, header=None, nrows=None) -> pd.DataFrame:
    return pd.read_excel(path, header=header, nrows=nrows)


def load_header_aliases(template_dir: Path) -> Dict[str, List[str]]:
    aliases = {key: list(values) for key, values in DEFAULT_ALIASES.items()}
    path = template_dir / "表头匹配对应表.xlsx"
    if not path.exists():
        return aliases
    df = read_excel_any(path, header=None)
    for _, row in df.iterrows():
        standard = normalize_text(row.get(0))
        raw = normalize_text(row.get(1))
        if not standard or not raw or standard in {"表头匹配对照表", "泾都药店名称"}:
            continue
        parts = [item for item in re.split(r"[=,，、|/]+", raw) if normalize_text(item)]
        aliases.setdefault(standard, [])
        for item in [standard, *parts]:
            if item not in aliases[standard]:
                aliases[standard].append(item)
    return aliases


def detect_header_row(raw: pd.DataFrame, aliases: Dict[str, List[str]]) -> int:
    alias_keys = {normalize_key(item) for values in aliases.values() for item in values}
    best_row = 0
    best_score = -1
    for idx in range(min(len(raw), 15)):
        values = [normalize_key(v) for v in raw.iloc[idx].tolist()]
        score = sum(1 for value in values if value in alias_keys)
        non_empty = sum(1 for value in values if value)
        if score > best_score or (score == best_score and non_empty > 2):
            best_row = idx
            best_score = score
    return best_row


def build_column_map(columns: Iterable[object], aliases: Dict[str, List[str]]) -> Dict[str, str]:
    normalized_to_column = {normalize_key(col): col for col in columns}
    result: Dict[str, str] = {}
    for standard, names in aliases.items():
        for name in names:
            key = normalize_key(name)
            if key in normalized_to_column:
                result[standard] = normalized_to_column[key]
                break
    return result


def load_supplier_header_patterns(template_dir: Path) -> Dict[str, set[str]]:
    path = template_dir / "表头反推供应商.xls"
    patterns: Dict[str, set[str]] = {}
    if not path.exists():
        return patterns
    df = read_excel_any(path, header=None)
    for _, row in df.iterrows():
        supplier = normalize_text(row.get(0))
        if not supplier or supplier in {"表头反推供应商", "供应商名称"}:
            continue
        headers = {normalize_key(v) for v in row.tolist()[1:] if normalize_text(v)}
        if headers:
            patterns[supplier] = headers
    return patterns


def load_map(path: Path, key_col: str, value_cols: List[str]) -> Dict[str, dict]:
    if not path.exists():
        return {}
    df = read_excel_any(path, header=0).fillna("")
    result = {}
    for _, row in df.iterrows():
        key = normalize_text(row.get(key_col))
        if not key:
            continue
        result[normalize_key(key)] = {col: normalize_text(row.get(col)) for col in value_cols}
        result[normalize_key(key)]["原名称"] = key
    return result


def load_supplier_codes(template_dir: Path) -> Dict[str, dict]:
    return load_map(template_dir / "供货商名称对应供货商名称编码.xls", "供货商名称", ["供货商名称编码"])


def load_supply_unit_codes(template_dir: Path) -> Dict[str, dict]:
    return load_map(
        template_dir / "供货单位名称对应供货单位名称编码对应采购人编码.xls",
        "供货单位名称",
        ["供货单位名称编码", "采购人编码"],
    )


def load_multipliers(template_dir: Path) -> Dict[str, Decimal]:
    path = template_dir / "特定供货商价格倍率.xls"
    if not path.exists():
        return {}
    df = read_excel_any(path, header=0).fillna("")
    result = {}
    for _, row in df.iterrows():
        supplier = normalize_text(row.get("供货商名称"))
        factor = to_decimal(row.get("倍率"))
        if supplier and factor:
            result[normalize_key(supplier)] = factor
    return result


def load_price_adjustments(template_dir: Path) -> List[dict]:
    path = template_dir / "药品价格上调明细表.xlsx"
    if not path.exists():
        return []
    df = read_excel_any(path, header=0).fillna("")
    rows = []
    for _, row in df.iterrows():
        rows.append(
            {
                "drug_code": normalize_text(row.get("药品编码")),
                "maker": normalize_key(row.get("厂家名称（和厂家发来的出库单一致）")),
                "type": normalize_text(row.get("上调类型（填比例或者固定金额）")),
                "adjusted_price": to_decimal(row.get("上调后的价格")),
                "value": to_decimal(row.get("上调值/比例")),
                "drug_name": normalize_key(row.get("药品名称")),
                "supplier": normalize_key(row.get("供货方名称")),
            }
        )
    return rows


def load_drug_catalog(template_dir: Path) -> List[dict]:
    path = template_dir / "药品资料6.3.xls"
    if not path.exists():
        return []
    df = read_excel_any(path, header=1).fillna("")
    rows = []
    for _, row in df.iterrows():
        code = normalize_text(row.get("药品编码"))
        name = normalize_text(row.get("药品名称"))
        approval = normalize_approval(row.get("批准文号"))
        if not code or not name:
            continue
        rows.append(
            {
                "code": code,
                "name": name,
                "name_key": normalize_key(name),
                "spec": normalize_text(row.get("规格")),
                "spec_key": normalize_key(row.get("规格")),
                "maker": normalize_text(row.get("生产厂家")),
                "maker_key": normalize_key(row.get("生产厂家")),
                "approval": approval,
                "unit": normalize_text(row.get("单位")),
                "unit_key": normalize_key(row.get("单位")),
            }
        )
    return rows


def find_name_by_fuzzy(mapping: Dict[str, dict], text: str) -> Tuple[str, dict]:
    key = normalize_key(text)
    if key in mapping:
        return mapping[key].get("原名称", text), mapping[key]
    for map_key, info in mapping.items():
        if key and (key in map_key or map_key in key):
            return info.get("原名称", text), info
    return text, {}


def infer_supplier(file_name: str, columns: Iterable[object], patterns: Dict[str, set[str]], supplier_codes: Dict[str, dict]) -> str:
    stem_key = normalize_key(Path(file_name).stem)
    for _, info in supplier_codes.items():
        name = info.get("原名称", "")
        key = normalize_key(name)
        if key and (key in stem_key or stem_key in key):
            return name
    col_keys = {normalize_key(col) for col in columns}
    best_supplier = ""
    best_score = 0
    for supplier, headers in patterns.items():
        score = len(col_keys & headers)
        if score > best_score:
            best_supplier = supplier
            best_score = score
    return best_supplier


def match_drug_catalog(row: dict, catalog: List[dict]) -> Tuple[Optional[dict], str, str, str]:
    approval = normalize_approval(row.get("批准文号"))
    spec_key = normalize_key(row.get("规格"))
    unit_key = normalize_key(row.get("单位"))
    name_key = normalize_key(row.get("药品名称"))
    if not approval:
        return None, "", "", "出库单批准文号为空或无有效编号，无法按批准文号对应药品基本资料"
    candidates = [item for item in catalog if item["approval"] == approval]
    if not candidates:
        return None, "", "", f"药品基本资料中未找到批准文号：{approval}"
    if len(candidates) == 1:
        return candidates[0], candidates[0]["spec"], candidates[0]["unit"], ""
    scored = []
    for item in candidates:
        score = 0
        if spec_key and spec_key == item["spec_key"]:
            score += 4
        elif spec_key and (spec_key in item["spec_key"] or item["spec_key"] in spec_key):
            score += 2
        if unit_key and unit_key == item["unit_key"]:
            score += 2
        if name_key and (name_key in item["name_key"] or item["name_key"] in name_key):
            score += 1
        scored.append((score, item))
    scored.sort(key=lambda pair: pair[0], reverse=True)
    if scored and scored[0][0] > 0:
        item = scored[0][1]
        return item, ", ".join(sorted({x["spec"] for x in candidates if x["spec"]})), ", ".join(sorted({x["unit"] for x in candidates if x["unit"]})), ""
    return None, ", ".join(sorted({x["spec"] for x in candidates if x["spec"]})), ", ".join(sorted({x["unit"] for x in candidates if x["unit"]})), "批准文号命中多条药品资料，但规格/单位/名称未能确定唯一药品"


def adjusted_price(
    supplier: str,
    unit_price: Optional[Decimal],
    drug: Optional[dict],
    row: dict,
    multipliers: Dict[str, Decimal],
    price_adjustments: List[dict],
) -> Tuple[Optional[Decimal], str, str, str]:
    if unit_price is None:
        return None, "", "", ""
    supplier_key = normalize_key(supplier)
    maker_key = normalize_key(row.get("生产厂家"))
    name_key = normalize_key(row.get("药品名称"))
    for item in price_adjustments:
        supplier_ok = not item["supplier"] or item["supplier"] in supplier_key or supplier_key in item["supplier"]
        maker_ok = not item["maker"] or item["maker"] in maker_key or maker_key in item["maker"]
        name_ok = not item["drug_name"] or item["drug_name"] in name_key or name_key in item["drug_name"]
        code_ok = drug and item["drug_code"] and item["drug_code"] == drug["code"]
        if supplier_ok and maker_ok and (name_ok or code_ok) and item["adjusted_price"] is not None:
            return item["adjusted_price"], item["type"], str(item.get("value") or ""), item.get("drug_code") or ""
    for key, factor in multipliers.items():
        if key and (key in supplier_key or supplier_key in key):
            value = (unit_price * factor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            return value, "供应商倍率", str(factor), drug["code"] if drug else ""
    return unit_price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "", "", drug["code"] if drug else ""


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, col in enumerate(ws.columns, start=1):
        max_len = 8
        for cell in col:
            max_len = max(max_len, min(len(normalize_text(cell.value)), 48))
        ws.column_dimensions[get_column_letter(idx)].width = max_len + 2


def workbook_bytes(sheets: Dict[str, List[dict]], columns: Dict[str, List[str]]) -> bytes:
    wb = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_name[:31]
        headers = columns[sheet_name]
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header, "") for header in headers])
        style_sheet(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def clean_files(upload_paths: List[Path], template_dir: Path) -> Tuple[bytes, List[CleanResult]]:
    aliases = load_header_aliases(template_dir)
    supplier_patterns = load_supplier_header_patterns(template_dir)
    supplier_codes = load_supplier_codes(template_dir)
    supply_unit_codes = load_supply_unit_codes(template_dir)
    multipliers = load_multipliers(template_dir)
    price_adjustments = load_price_adjustments(template_dir)
    catalog = load_drug_catalog(template_dir)

    standard_rows: List[dict] = []
    failure_rows: List[dict] = []
    failure_text_blocks: List[str] = []
    summaries: List[CleanResult] = []

    for path in upload_paths:
        raw = read_excel_any(path, header=None)
        header_row = detect_header_row(raw, aliases)
        df = read_excel_any(path, header=header_row).fillna("")
        col_map = build_column_map(df.columns, aliases)
        supplier = infer_supplier(path.name, df.columns, supplier_patterns, supplier_codes)
        supplier, supplier_info = find_name_by_fuzzy(supplier_codes, supplier)
        supplier_code = supplier_info.get("供货商名称编码", "")

        total = success = failed = 0
        for excel_row_no, (_, src) in enumerate(df.iterrows(), start=header_row + 2):
            row = {field: normalize_text(src.get(col_map[field])) if field in col_map else "" for field in DEFAULT_ALIASES}
            if not (row.get("药品名称") or row.get("批准文号") or row.get("批号")):
                continue
            total += 1

            supply_unit_name, supply_unit_info = find_name_by_fuzzy(supply_unit_codes, row.get("供货单位名称"))
            unit_price = to_decimal(row.get("单价"))
            drug, candidate_specs, candidate_units, fail_reason = match_drug_catalog(row, catalog)
            adjusted, adjust_type, adjust_value, adjust_code = adjusted_price(
                supplier, unit_price, drug, row, multipliers, price_adjustments
            )

            standard_row = {col: "" for col in STANDARD_COLUMNS}
            standard_row.update(
                {
                    "批准文号": row.get("批准文号"),
                    "药品名称": row.get("药品名称"),
                    "最终药品编码": drug["code"] if drug else "",
                    "批号": row.get("批号"),
                    "生产日期": row.get("生产日期"),
                    "有效期至": row.get("有效期至"),
                    "数量": row.get("数量"),
                    "上调后单价": float(adjusted) if adjusted is not None else "",
                    "供货商名称": supplier,
                    "供货商名称编码": supplier_code,
                    "供货单位名称": supply_unit_name,
                    "供货单位名称编码": supply_unit_info.get("供货单位名称编码", ""),
                    "采购人编码": supply_unit_info.get("采购人编码", ""),
                    "生产厂家": row.get("生产厂家"),
                    "单价": float(unit_price) if unit_price is not None else "",
                    "规格": row.get("规格"),
                    "厂家批准文号清洗": normalize_approval(row.get("批准文号")),
                    "厂家商品规格清洗": row.get("规格"),
                    "单位": row.get("单位"),
                    "厂家包装单位清洗": row.get("单位"),
                    "单价上调类型": adjust_type,
                    "单价上调值/比例": adjust_value,
                    "单价上调表中的药品编码": adjust_code,
                    "候选规格": candidate_specs,
                    "候选单位": candidate_units,
                }
            )
            standard_rows.append(standard_row)

            if fail_reason:
                failed += 1
                failure = {
                    "源文件": path.name,
                    "标准表行号": excel_row_no,
                    "供货商名称": supplier,
                    "药品名称": row.get("药品名称"),
                    "批准文号": row.get("批准文号"),
                    "规格": row.get("规格"),
                    "单位": row.get("单位"),
                    "批号": row.get("批号"),
                    "失败阶段": "approval",
                    "失败原因": fail_reason,
                }
                failure_rows.append(failure)
                failure_text_blocks.append(
                    "\n".join(
                        [
                            f"[{len(failure_rows)}] 标准表行号：{excel_row_no}",
                            f"源文件：{path.name}",
                            f"药品名称：{row.get('药品名称')}",
                            f"批准文号：{row.get('批准文号')}",
                            f"规格：{row.get('规格')}",
                            f"单位：{row.get('单位')}",
                            f"批号：{row.get('批号')}",
                            "失败阶段：approval",
                            f"失败原因：{fail_reason}",
                        ]
                    )
                )
            else:
                success += 1

        summaries.append(CleanResult(path.name, supplier, total, success, failed))

    result_xlsx = workbook_bytes(
        {"标准化清洗结果": standard_rows, "失败明细": failure_rows},
        {"标准化清洗结果": STANDARD_COLUMNS, "失败明细": ["源文件", "标准表行号", "供货商名称", "药品名称", "批准文号", "规格", "单位", "批号", "失败阶段", "失败原因"]},
    )

    summary_rows = [
        {
            "文件名": item.file_name,
            "识别供货商": item.supplier,
            "总行数": item.total_rows,
            "成功行数": item.success_rows,
            "失败行数": item.failed_rows,
        }
        for item in summaries
    ]
    summary_xlsx = workbook_bytes(
        {"清洗摘要": summary_rows},
        {"清洗摘要": ["文件名", "识别供货商", "总行数", "成功行数", "失败行数"]},
    )

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("标准化清洗结果.xlsx", result_xlsx)
        zf.writestr("清洗摘要.xlsx", summary_xlsx)
        zf.writestr("失败原因.txt", "\n\n".join(failure_text_blocks) if failure_text_blocks else "本次没有失败记录。")
    return zip_buf.getvalue(), summaries
