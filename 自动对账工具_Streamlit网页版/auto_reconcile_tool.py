"""
泾都药业自动对账工具

用途：
1. 读取药店数据、银行数据、申报表数据。
2. 完成“药店数据 vs 银行货款支出”的自动对账。
3. 预留“医保收款 vs 申报表”的流程和 AI 接口位置，方便后续补充规则。
4. 导出带筛选、冻结窗格、颜色标记的 Excel 结果表。

重要设计：
- 不写死任何绝对路径。
- 默认读取程序同级目录下的“对账数据表准备”文件夹。
- 也支持在界面中手动选择文件。
- 打包成 EXE 后，所有路径都以 EXE 所在目录为基准。
"""

from __future__ import annotations

import datetime as _dt
import difflib
import json
import os
import re
import shutil
import sys
import traceback
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Tuple

import pandas as pd
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:
    tk = None
    filedialog = messagebox = ttk = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


APP_TITLE = "泾都药业自动对账工具"
DEFAULT_DATA_FOLDER_NAME = "对账数据表准备"
OUTPUT_PREFIX = "对账结果汇总表"

# 默认文件名。后续如果文件名变化，可以直接改这里。
DEFAULT_DRUG_FILE = "药店数据模板.xls"
DEFAULT_BANK_FILE = "银行数据模板.xls"
DEFAULT_DECLARE_FILE = "申报表模板.xlsx"
HISTORY_FOLDER_NAME = "历史申报表库"
TEMPLATE_FOLDER_NAME = "模板文件"
DEFAULT_CONFIG_FILE = "业务规则配置总表.xlsx"
SETTINGS_FILE_NAME = "工具设置.json"

# 金额极小差异阈值：不自动算成功，只在备注里提示“差异很小”。
TINY_DIFF_THRESHOLD = Decimal("0.10")
HAITONG_COMPANY = "陕西海通药业有限责任公司"


class UserVisibleError(Exception):
    """用于向用户弹窗展示的业务错误。"""


@dataclass
class FileSelection:
    """界面中当前选择的三个文件。"""

    drug_file: Optional[Path] = None
    bank_file: Optional[Path] = None
    declare_file: Optional[Path] = None
    data_folder: Optional[Path] = None
    output_folder: Optional[Path] = None
    config_file: Optional[Path] = None
    high_invoice_file: Optional[Path] = None
    declare_unit: str = ""
    bank_account_name: str = ""
    unit_check_message: str = ""


@dataclass
class ReconcileRow:
    """药店和银行对账后的单行结果。"""

    category: str
    result: str
    match_type: str
    ticket_no: str
    bank_code: str
    drug_amount: Optional[Decimal]
    bank_amount: Optional[Decimal]
    diff: Optional[Decimal]
    note: str
    bank_original_amount: Optional[Decimal] = None
    bank_adjusted_amount: Optional[Decimal] = None
    adjustment_rule: str = ""
    drug_rows: str = ""
    bank_rows: str = ""
    supplier: str = ""
    bank_counterparty: str = ""
    bank_purpose: str = ""
    bank_time: str = ""


def get_base_dir() -> Path:
    """
    获取程序基准目录。

    - 普通 Python 运行：源码所在目录。
    - PyInstaller 打包后：EXE 所在目录。
    """

    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def normalize_text(value: object) -> str:
    """把单元格内容清洗成普通字符串。"""

    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    # 去掉银行备注里常见的句号、空格和 XML 标签边界对匹配的干扰。
    return text.replace("\u3000", " ").strip()


def safe_path_name(text: str) -> str:
    """把单位名称转换成可用于 Windows 文件夹名的安全文本。"""

    text = normalize_text(text) or "未知单位"
    return re.sub(r'[<>:"/\\|?*]+', "_", text).strip(" .")


def to_decimal(value: object, field_name: str = "金额") -> Decimal:
    """
    把 Excel 单元格金额转成 Decimal，避免浮点数比较产生误差。

    空值按 0 处理；无法转换时抛出清晰错误。
    """

    if value is None or value == "":
        return Decimal("0.00")
    try:
        if pd.isna(value):
            return Decimal("0.00")
    except (TypeError, ValueError):
        pass
    try:
        text = str(value).replace(",", "").strip()
        if not text or text.lower() in {"nan", "none", "nat"}:
            return Decimal("0.00")
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise UserVisibleError(f"{field_name}格式错误，无法识别为数字：{value}") from exc


def get_default_config_path() -> Path:
    """默认业务规则配置总表路径。"""

    return get_base_dir() / TEMPLATE_FOLDER_NAME / DEFAULT_CONFIG_FILE


def get_settings_path() -> Path:
    """工具设置文件路径，用于记住上次选择的文件。"""

    return get_base_dir() / SETTINGS_FILE_NAME


def read_settings() -> dict:
    """读取上次选择的路径设置。"""

    path = get_settings_path()
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_settings(selection: FileSelection) -> None:
    """保存当前选择路径，供下次启动自动恢复。"""

    data = {
        "drug_file": str(selection.drug_file or ""),
        "bank_file": str(selection.bank_file or ""),
        "declare_file": str(selection.declare_file or ""),
        "output_folder": str(selection.output_folder or ""),
        "config_file": str(selection.config_file or ""),
        "high_invoice_file": str(selection.high_invoice_file or ""),
    }
    try:
        get_settings_path().write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def load_settings_into_selection(selection: FileSelection) -> None:
    """把历史路径恢复到 FileSelection。"""

    data = read_settings()
    for key in ["drug_file", "bank_file", "declare_file", "config_file", "high_invoice_file"]:
        value = data.get(key)
        if value:
            setattr(selection, key, Path(value))
    if data.get("output_folder"):
        selection.output_folder = Path(data["output_folder"])


def clear_main_file_settings(selection: FileSelection) -> None:
    """清空首页四个路径，不清空配置路径。"""

    selection.drug_file = None
    selection.bank_file = None
    selection.declare_file = None
    selection.data_folder = None
    selection.output_folder = None
    save_settings(selection)


def path_from_text(text: str) -> Optional[Path]:
    """把界面输入框里的路径转成 Path；空白则返回 None。"""

    text = normalize_text(text)
    return Path(text) if text else None


def fuzzy_score(a: str, b: str) -> float:
    """公司名称模糊匹配分数。"""

    a = normalize_text(a)
    b = normalize_text(b)
    if not a or not b:
        return 0.0
    if a in b or b in a:
        return 1.0
    return difflib.SequenceMatcher(None, a, b).ratio()


def company_core(text: str) -> str:
    """提取公司名称中更有区分度的核心部分，降低通用后缀误匹配。"""

    value = normalize_text(text)
    value = re.sub(r"[（(].*?[）)]", "", value)
    for word in [
        "有限责任公司",
        "股份有限公司",
        "集团有限公司",
        "有限公司",
        "有限责任",
        "股份",
        "集团",
        "医药",
        "药业",
        "科技",
        "商贸",
        "公司",
        "有限",
    ]:
        value = value.replace(word, "")
    return re.sub(r"\s+", "", value)


def company_matches(company: str, text: str, threshold: float = 0.55) -> bool:
    """判断公司名称是否能与银行户名/用途模糊匹配。"""

    company = normalize_text(company)
    text = normalize_text(text)
    if not company or not text:
        return False
    if company in text or text in company:
        return True

    company_key = company_core(company)
    text_key = company_core(text)
    if not company_key or not text_key:
        return False
    if len(company_key) >= 4 and len(text_key) >= 4 and (company_key in text_key or text_key in company_key):
        return True
    return difflib.SequenceMatcher(None, company_key, text_key).ratio() >= max(threshold, 0.78)


def special_company_matches(rule_company: str, company: str) -> bool:
    """金额系数这类强规则必须严格命中，不能用宽松公司名模糊匹配。"""

    rule = normalize_text(rule_company)
    target = normalize_text(company)
    if not rule or not target:
        return False
    if rule in target or target in rule:
        return True
    rule_key = company_core(rule)
    target_key = company_core(target)
    if rule_key and target_key and (rule_key in target_key or target_key in rule_key):
        return True
    return "海通" in rule and "海通" in target


def load_business_config(config_file: Optional[Path] = None) -> dict:
    """
    读取业务规则配置。

    用户上传配置时优先使用上传文件；没有上传时读取默认配置总表；
    默认配置也不存在时使用代码内置规则。
    """

    config = {
        "special_multipliers": {HAITONG_COMPANY: Decimal("1.22")},
        "monthly_companies": [],
        "medical_discount_items": {
            "定点药店购药-个人账户-职工": Decimal("0.95"),
            "普通门诊-基本医保-职工": Decimal("0.95"),
            "门诊慢特病-基本医保-居民": Decimal("0.95"),
            "门诊慢特病-基本医保-职工": Decimal("0.95"),
        },
        "fuzzy_threshold": Decimal("0.55"),
    }

    path = config_file if config_file and config_file.exists() else get_default_config_path()
    if not path.exists():
        return config

    try:
        wb = load_workbook(path, data_only=True, read_only=True)

        if "特殊金额系数" in wb.sheetnames:
            ws = wb["特殊金额系数"]
            config["special_multipliers"] = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                company = normalize_text(row[0] if len(row) > 0 else "")
                multiplier = row[1] if len(row) > 1 else ""
                enabled = normalize_text(row[2] if len(row) > 2 else "是")
                if company and enabled != "否":
                    config["special_multipliers"][company] = to_decimal(multiplier or 1, "特殊金额系数")

        if "银行1笔付款公司" in wb.sheetnames:
            ws = wb["银行1笔付款公司"]
            companies = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                company = normalize_text(row[0] if len(row) > 0 else "")
                enabled = normalize_text(row[1] if len(row) > 1 else "是")
                if company and enabled != "否":
                    companies.append(company)
            config["monthly_companies"] = companies

        if "医保折扣规则" in wb.sheetnames:
            ws = wb["医保折扣规则"]
            discount_items = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = normalize_text(row[0] if len(row) > 0 else "")
                factor = row[1] if len(row) > 1 else ""
                enabled = normalize_text(row[2] if len(row) > 2 else "是")
                if item and enabled != "否":
                    discount_items[item] = to_decimal(factor or 1, "医保折扣系数")
            config["medical_discount_items"] = discount_items

        if "基础参数" in wb.sheetnames:
            ws = wb["基础参数"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                key = normalize_text(row[0] if len(row) > 0 else "")
                value = row[1] if len(row) > 1 else ""
                if key == "公司名称模糊匹配阈值":
                    config["fuzzy_threshold"] = to_decimal(value or "0.55", "公司名称模糊匹配阈值")
    except Exception as exc:
        raise UserVisibleError(f"读取业务规则配置失败：{path.name}。原因：{exc}") from exc

    return config


def read_excel_any(path: Path, header: Optional[int] = None) -> pd.DataFrame:
    """
    宽松读取 Excel。

    说明：
    - 银行表是真正老版 .xls，需要 xlrd。
    - 药店表虽然后缀是 .xls，但内容可能是新版 xlsx，需要 openpyxl。
    - 这里根据文件内容自动选择读取方式，而不是只看后缀。
    """

    if not path.exists():
        raise UserVisibleError(f"文件不存在：{path.name}")

    try:
        with path.open("rb") as f:
            signature = f.read(8)

        # PK 开头代表新版 Office 压缩格式，即使后缀写成 .xls，也按 openpyxl 读取。
        if signature.startswith(b"PK"):
            return pd.read_excel(path, header=header, engine="openpyxl")

        # D0 CF 11 E0 是老版 Office 二进制格式，用 xlrd 直接读取。
        # 不经 pandas.read_excel，是为了避免某些环境里 pandas 对 xlrd 版本元数据检查失败。
        if signature.startswith(b"\xd0\xcf\x11\xe0"):
            import xlrd

            book = xlrd.open_workbook(str(path))
            sheet = book.sheet_by_index(0)
            rows = [sheet.row_values(row_index) for row_index in range(sheet.nrows)]
            if header is None:
                return pd.DataFrame(rows)
            if header >= len(rows):
                raise UserVisibleError(f"读取 Excel 失败：{path.name}。标题行超出表格范围。")
            columns = [normalize_text(col) for col in rows[header]]
            return pd.DataFrame(rows[header + 1 :], columns=columns)

        # 兜底：让 pandas 自己判断。
        return pd.read_excel(path, header=header)
    except ImportError as exc:
        raise UserVisibleError(
            "缺少读取老版 .xls 文件所需依赖 xlrd。请按说明重新打包，确保 xlrd 被包含。"
        ) from exc
    except Exception as exc:
        raise UserVisibleError(f"读取 Excel 失败：{path.name}。原因：{exc}") from exc


def require_columns(df: pd.DataFrame, required: Iterable[str], file_label: str) -> None:
    """检查表格是否包含必要字段。"""

    missing = [col for col in required if col not in df.columns]
    if missing:
        raise UserVisibleError(f"{file_label}缺少必要字段：{', '.join(missing)}")


def extract_bank_code(purpose: str) -> str:
    """
    从银行“交易用途”里提取“货款”后面的字母数字编号。

    例：
    - 货款XXD202605290818304908. -> XXD202605290818304908
    - 货款. -> 空字符串，表示员工未填写编号，不参与匹配。
    """

    text = normalize_text(purpose)
    match = re.search(r"货款\s*([A-Za-z0-9]+)", text)
    if not match:
        return ""
    return match.group(1).strip()


def code_similarity(a: str, b: str) -> float:
    """
    计算两个票据号的简易相似度。

    规则偏业务可解释：
    - 完全相同：1.0
    - 一个是另一个的前缀/包含：较高分
    - 其他情况：使用公共前缀占比作为参考

    这里只用于“疑似匹配”，不会替代完全匹配。
    """

    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    if a in b or b in a:
        return min(len(a), len(b)) / max(len(a), len(b))

    common = 0
    for ca, cb in zip(a, b):
        if ca != cb:
            break
        common += 1
    return common / max(len(a), len(b))


def find_fuzzy_match(ticket_no: str, bank_candidates: Dict[str, dict]) -> Optional[str]:
    """
    为药店票据号寻找一个银行疑似编号。

    当前策略：
    - 首先允许一个编号包含另一个编号；
    - 相似度达到 0.70 以上才接受；
    - 如果多个候选，取相似度最高的。

    后续如果你想更严格或更宽松，只改这里即可。
    """

    best_code = None
    best_score = Decimal("0")
    for code in bank_candidates:
        score = Decimal(str(code_similarity(ticket_no, code)))
        if score > best_score:
            best_score = score
            best_code = code

    if best_code and best_score >= Decimal("0.70"):
        return best_code
    return None


def load_drug_data(path: Path) -> Dict[str, dict]:
    """
    读取药店数据，并按“票据号码”汇总“应结算”金额。

    同一个票据号码多行时：
    - 金额累加；
    - 原始行号合并记录；
    - 供货商合并去重。
    """

    df = read_excel_any(path, header=1)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["票据号码", "应结算", "供货商"], "药店数据表")

    grouped: Dict[str, dict] = {}
    for index, row in df.iterrows():
        ticket_no = normalize_text(row.get("票据号码"))
        if not ticket_no:
            continue

        amount = to_decimal(row.get("应结算"), "药店应结算")
        excel_row = index + 3  # header=1 表示 Excel 第 2 行是标题，数据从第 3 行开始。
        supplier = normalize_text(row.get("供货商"))

        if ticket_no not in grouped:
            grouped[ticket_no] = {
                "ticket_no": ticket_no,
                "amount": Decimal("0.00"),
                "rows": [],
                "suppliers": set(),
            }
        grouped[ticket_no]["amount"] += amount
        grouped[ticket_no]["rows"].append(str(excel_row))
        if supplier:
            grouped[ticket_no]["suppliers"].add(supplier)

    return grouped


def load_drug_records(path: Path) -> List[dict]:
    """读取药店明细记录，供多规则对账使用。"""

    df = read_excel_any(path, header=1)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["票据号码", "应结算", "供货商"], "药店数据表")

    records = []
    for index, row in df.iterrows():
        supplier = normalize_text(row.get("供货商"))
        amount = to_decimal(row.get("应结算"), "药店应结算")
        ticket_no = normalize_text(row.get("票据号码"))
        if not supplier and amount == Decimal("0.00") and not ticket_no:
            continue
        records.append(
            {
                "row": index + 3,
                "ticket_no": ticket_no,
                "amount": amount,
                "supplier": supplier,
                "department": normalize_text(row.get("部门")),
                "date": normalize_text(row.get("单据日期")),
                "used": False,
            }
        )
    return records


def load_bank_data(path: Path) -> Tuple[Dict[str, dict], List[dict]]:
    """
    读取银行数据。

    返回：
    - 带货款编号的记录，按编号汇总支出金额；
    - 没有编号的货款记录，仅记录到运行摘要，不参与匹配。
    """

    df = read_excel_any(path, header=2)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["交易时间", "收入金额", "支出金额", "对方户名", "交易用途"], "银行数据表")

    grouped: Dict[str, dict] = {}
    no_code_rows: List[dict] = []

    for index, row in df.iterrows():
        purpose = normalize_text(row.get("交易用途"))
        if "货款" not in purpose:
            continue

        bank_code = extract_bank_code(purpose)
        expense = to_decimal(row.get("支出金额"), "银行支出金额")
        excel_row = index + 4  # header=2 表示 Excel 第 3 行是标题，数据从第 4 行开始。
        counterparty = normalize_text(row.get("对方户名"))
        bank_time = normalize_text(row.get("交易时间"))

        if not bank_code:
            no_code_rows.append(
                {
                    "row": excel_row,
                    "amount": expense,
                    "counterparty": counterparty,
                    "purpose": purpose,
                    "bank_time": bank_time,
                }
            )
            continue

        if bank_code not in grouped:
            grouped[bank_code] = {
                "bank_code": bank_code,
                "amount": Decimal("0.00"),
                "rows": [],
                "counterparties": set(),
                "purposes": [],
                "times": [],
            }
        grouped[bank_code]["amount"] += expense
        grouped[bank_code]["rows"].append(str(excel_row))
        if counterparty:
            grouped[bank_code]["counterparties"].add(counterparty)
        if purpose:
            grouped[bank_code]["purposes"].append(purpose)
        if bank_time:
            grouped[bank_code]["times"].append(bank_time)

    return grouped, no_code_rows


def load_bank_records(path: Path) -> List[dict]:
    """读取银行支出明细，保留带编号和不带编号的货款/公司付款记录。"""

    df = read_excel_any(path, header=2)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["交易时间", "收入金额", "支出金额", "对方户名", "交易用途"], "银行数据表")
    records = []
    for index, row in df.iterrows():
        try:
            expense = to_decimal(row.get("支出金额"), "银行支出金额")
        except UserVisibleError:
            summary_text = " ".join(
                normalize_text(row.get(col)) for col in ("交易时间", "收入金额", "支出金额", "账户余额")
            )
            if "总收入" in summary_text or "总支出" in summary_text:
                continue
            raise
        if expense == Decimal("0.00"):
            continue
        purpose = normalize_text(row.get("交易用途"))
        counterparty = normalize_text(row.get("对方户名"))
        if not purpose and not counterparty:
            continue
        records.append(
            {
                "row": index + 4,
                "time": normalize_text(row.get("交易时间")),
                "amount": expense,
                "counterparty": counterparty,
                "purpose": purpose,
                "code": extract_bank_code(purpose),
                "used": False,
            }
        )
    return records


def parse_date(value: object) -> Optional[_dt.datetime]:
    """尽量把 Excel 日期/字符串转成 datetime。"""

    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value
    if isinstance(value, _dt.date):
        return _dt.datetime.combine(value, _dt.time())
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
        try:
            return _dt.datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def load_high_invoice_records(path: Optional[Path]) -> List[dict]:
    """读取高开票差额数据。"""

    if not path or not path.exists():
        return []
    df = read_excel_any(path, header=0)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["供货商名称", "日期", "合计金额"], "高开票对账表")
    records = []
    for index, row in df.iterrows():
        supplier = normalize_text(row.get("供货商名称"))
        if not supplier:
            continue
        records.append(
            {
                "row": index + 2,
                "supplier": supplier,
                "branch": normalize_text(row.get("分部")),
                "date": parse_date(row.get("日期")),
                "amount": to_decimal(row.get("合计金额"), "高开票合计金额"),
                "doc_no": normalize_text(row.get("单据号")),
                "used": False,
            }
        )
    return records


def compare_amounts(drug_amount: Decimal, bank_amount: Decimal) -> Tuple[str, Decimal, str]:
    """比较药店应结算和银行支出金额，返回结果、差额和备注。"""

    diff = (drug_amount - bank_amount).quantize(Decimal("0.01"))
    if diff == Decimal("0.00"):
        return "对账成功", diff, ""

    note = ""
    if abs(diff) <= TINY_DIFF_THRESHOLD:
        note = f"金额存在极小差异 {diff}，请人工确认"

    if diff > 0:
        return "药店金额大于银行金额", diff, note or "银行少付，或药店上货金额偏大"
    return "银行金额大于药店金额", diff, note or "银行多付，或药店上货金额偏小"


def multiplier_for_company(company: str, config: dict) -> Decimal:
    """根据公司名称获取银行金额系数。"""

    for name, factor in config.get("special_multipliers", {}).items():
        if special_company_matches(name, company):
            return factor
    return Decimal("1.00")


def adjusted_bank_amount(amount: Decimal, company: str, config: dict, extra: Decimal = Decimal("0.00")) -> Tuple[Decimal, str]:
    """计算银行对账金额，并返回调整规则。"""

    factor = multiplier_for_company(company, config)
    adjusted = (amount * factor + extra).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rules = []
    if factor != Decimal("1.00"):
        rules.append(f"{company}：银行金额×{factor}")
    if extra:
        rules.append(f"加高开票差额{extra}")
    return adjusted, "；".join(rules)


def extract_branch_token(text: object) -> str:
    """从户名/分部名称中提取“二分部、五分部”这类分部标识。"""

    value = normalize_text(text)
    match = re.search(r"([一二三四五六七八九十\d]+)分部", value)
    return match.group(0) if match else ""


def high_invoice_branch_matches(current_branch_text: str, high_invoice_branch: str) -> bool:
    """高开票必须明确属于当前分部，避免把其他分部的金额加进来。"""

    current_branch = extract_branch_token(current_branch_text)
    high_branch = extract_branch_token(high_invoice_branch)
    if not current_branch or not high_branch:
        return False
    return current_branch == high_branch


def record_date(record: dict, key: str) -> Optional[_dt.datetime]:
    return parse_date(record.get(key))


def days_between(a: Optional[_dt.datetime], b: Optional[_dt.datetime]) -> int:
    if not a or not b:
        return 9999
    return abs((a.date() - b.date()).days)


def is_special_multiplier_company(company: str, config: dict) -> bool:
    return multiplier_for_company(company, config) != Decimal("1.00")


def final_bank_amount(
    bank_amount: Decimal,
    company: str,
    config: dict,
    high_invoice_extra: Decimal = Decimal("0.00"),
    high_invoice_rows: Optional[List[dict]] = None,
) -> Tuple[Decimal, str]:
    """最终药店-银行对账口径：海通乘系数；其他公司银行金额加高开票。"""

    factor = multiplier_for_company(company, config)
    if factor != Decimal("1.00"):
        adjusted = (bank_amount * factor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return adjusted, f"{company}：固定规则，银行金额×{factor}"

    adjusted = (bank_amount + high_invoice_extra).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if high_invoice_extra:
        rows = ", ".join(str(item["row"]) for item in (high_invoice_rows or []))
        return adjusted, f"银行转账金额 + 高开票合计{high_invoice_extra}（高开票原始行：{rows}）"
    return adjusted, "银行转账金额"


def high_invoice_candidates(
    records: List[dict], company: str, current_branch_text: str, only_unused: bool = True
) -> List[dict]:
    return sorted(
        [
            item
            for item in records
            if (not only_unused or not item["used"])
            and company_matches(company, item["supplier"])
            and high_invoice_branch_matches(current_branch_text, item.get("branch", ""))
        ],
        key=lambda item: (item.get("date") or _dt.datetime.min, item["row"]),
    )


def take_high_invoice_for_bank(bank: dict, candidates: List[dict], fallback_index: int) -> Tuple[Decimal, List[dict]]:
    """按银行日期优先匹配高开票；日期都不近时按同公司同分部顺序兜底。"""

    available = [item for item in candidates if not item["used"]]
    if not available:
        return Decimal("0.00"), []

    bank_dt = record_date(bank, "time")
    dated = [item for item in available if item.get("date")]
    if bank_dt and dated:
        close = sorted(dated, key=lambda item: (days_between(bank_dt, item.get("date")), item["row"]))
        if close and days_between(bank_dt, close[0].get("date")) <= 3:
            close[0]["used"] = True
            return close[0]["amount"], [close[0]]

    if fallback_index < len(candidates):
        item = candidates[fallback_index]
        if not item["used"]:
            item["used"] = True
            return item["amount"], [item]

    return Decimal("0.00"), []


def take_monthly_high_invoices(company: str, bank_items: List[dict], high_invoice_records: List[dict], current_branch_text: str) -> Tuple[Decimal, List[dict]]:
    """银行1笔付款公司按同公司、同分部、同月/近日期汇总高开票。"""

    candidates = high_invoice_candidates(high_invoice_records, company, current_branch_text)
    if not candidates or not bank_items:
        return Decimal("0.00"), []

    bank_dates = [record_date(item, "time") for item in bank_items if record_date(item, "time")]
    used = []
    for item in candidates:
        hi_date = item.get("date")
        if not hi_date:
            continue
        same_month = any(bank_dt and bank_dt.year == hi_date.year and bank_dt.month == hi_date.month for bank_dt in bank_dates)
        near_bank = any(days_between(bank_dt, hi_date) <= 3 for bank_dt in bank_dates)
        if same_month or near_bank:
            item["used"] = True
            used.append(item)
    return summarize_records(used), used


def group_drug_by_supplier(records: List[dict], supplier: str, only_unused: bool = True) -> List[dict]:
    """按供货商模糊匹配药店记录。"""

    return [
        item
        for item in records
        if (not only_unused or not item["used"])
        and item["amount"] != Decimal("0.00")
        and company_matches(supplier, item["supplier"])
    ]


def find_bank_by_company(records: List[dict], company: str, only_unused: bool = True) -> List[dict]:
    """按公司名称在银行对方户名优先、交易用途兜底匹配银行记录。"""

    matched = []
    for item in records:
        if only_unused and item["used"]:
            continue
        if company_matches(company, item["counterparty"]) or company_matches(company, item["purpose"]):
            matched.append(item)
    return matched


def summarize_records(records: List[dict], amount_key: str = "amount") -> Decimal:
    """汇总金额。"""

    return sum((item[amount_key] for item in records), Decimal("0.00")).quantize(Decimal("0.01"))


def make_reconcile_row(
    category: str,
    result: str,
    match_type: str,
    ticket_no: str,
    bank_code: str,
    drug_amount: Optional[Decimal],
    bank_original: Optional[Decimal],
    bank_adjusted: Optional[Decimal],
    diff: Optional[Decimal],
    note: str,
    drug_items: List[dict],
    bank_items: List[dict],
    supplier: str,
    adjustment_rule: str = "",
) -> ReconcileRow:
    """统一构造药店银行对账结果行。"""

    return ReconcileRow(
        category=category,
        result=result,
        match_type=match_type,
        ticket_no=ticket_no,
        bank_code=bank_code,
        drug_amount=drug_amount,
        bank_amount=bank_adjusted,
        bank_original_amount=bank_original,
        bank_adjusted_amount=bank_adjusted,
        adjustment_rule=adjustment_rule,
        diff=diff,
        note=note,
        drug_rows=", ".join(str(item["row"]) for item in drug_items),
        bank_rows=", ".join(str(item["row"]) for item in bank_items),
        supplier=supplier or "；".join(sorted({item.get("supplier", "") for item in drug_items if item.get("supplier")})),
        bank_counterparty="；".join(sorted({item.get("counterparty", "") for item in bank_items if item.get("counterparty")})),
        bank_purpose="；".join(item.get("purpose", "") for item in bank_items if item.get("purpose")),
        bank_time="；".join(item.get("time", "") for item in bank_items if item.get("time")),
    )


def reconcile_drug_bank_v2(
    drug_records: List[dict],
    bank_records: List[dict],
    high_invoice_records: List[dict],
    config: dict,
) -> Tuple[List[ReconcileRow], dict, List[dict]]:
    """新版药店-银行对账：高开票、月汇总、普通票据号三段处理。"""

    results: List[ReconcileRow] = []
    threshold = float(config.get("fuzzy_threshold", Decimal("0.55")))

    # 1. 高开票差额对账：按公司 + 日期顺序匹配银行付款，银行金额 + 差额 后对账。
    for company in sorted({item["supplier"] for item in high_invoice_records}):
        hi_items = sorted([x for x in high_invoice_records if x["supplier"] == company], key=lambda x: x["date"] or _dt.datetime.min)
        drug_items = sorted(group_drug_by_supplier(drug_records, company), key=lambda x: parse_date(x.get("date")) or _dt.datetime.min)
        bank_items = sorted(find_bank_by_company(bank_records, company), key=lambda x: parse_date(x.get("time")) or _dt.datetime.min)
        pair_count = max(len(hi_items), len(bank_items), 1)
        for idx in range(pair_count):
            hi = hi_items[idx] if idx < len(hi_items) else None
            bank = bank_items[idx] if idx < len(bank_items) else None
            related_drugs = [drug_items[idx]] if idx < len(drug_items) else []
            if len(drug_items) != len(hi_items) and idx == 0:
                related_drugs = drug_items
            drug_amount = summarize_records(related_drugs) if related_drugs else Decimal("0.00")
            bank_original = bank["amount"] if bank else None
            extra = hi["amount"] if hi else Decimal("0.00")
            bank_adjusted, rule = adjusted_bank_amount(bank_original or Decimal("0.00"), company, config, extra)
            result, diff, note = compare_amounts(drug_amount, bank_adjusted)
            if not bank:
                result, diff, note = "银行未找到", None, "高开票公司在银行中未找到对应付款"
                bank_adjusted = None
            if not related_drugs:
                result, diff, note = "药店未找到", None, "高开票公司在药店数据中未找到对应入库"
            if hi:
                hi["used"] = True
            if bank:
                bank["used"] = True
            for item in related_drugs:
                item["used"] = True
            results.append(
                make_reconcile_row(
                    "高开票差额对账",
                    result,
                    "公司+日期顺序",
                    "高开票汇总",
                    "",
                    drug_amount,
                    bank_original,
                    bank_adjusted,
                    diff,
                    note,
                    related_drugs,
                    [bank] if bank else [],
                    company,
                    rule,
                )
            )

    # 2. 银行1笔付款公司：按公司月汇总对账。
    for company in config.get("monthly_companies", []):
        drug_items = group_drug_by_supplier(drug_records, company)
        bank_items = find_bank_by_company(bank_records, company)
        if not drug_items and not bank_items:
            continue
        drug_amount = summarize_records(drug_items)
        bank_original = summarize_records(bank_items)
        bank_adjusted, rule = adjusted_bank_amount(bank_original, company, config)
        result, diff, note = compare_amounts(drug_amount, bank_adjusted)
        for item in drug_items:
            item["used"] = True
        for item in bank_items:
            item["used"] = True
        results.append(
            make_reconcile_row(
                "银行1笔付款公司汇总",
                result,
                "公司月汇总",
                "公司月汇总",
                "",
                drug_amount,
                bank_original,
                bank_adjusted,
                diff,
                note,
                drug_items,
                bank_items,
                company,
                rule,
            )
        )

    # 3. 普通票据号对账。
    drug_group: Dict[str, dict] = {}
    for item in drug_records:
        if item["used"] or not item["ticket_no"]:
            continue
        group = drug_group.setdefault(item["ticket_no"], {"items": [], "amount": Decimal("0.00"), "suppliers": set()})
        group["items"].append(item)
        group["amount"] += item["amount"]
        if item["supplier"]:
            group["suppliers"].add(item["supplier"])

    bank_by_code: Dict[str, dict] = {}
    no_code_rows = []
    for item in bank_records:
        if item["used"]:
            continue
        if not item["code"]:
            no_code_rows.append(item)
            continue
        group = bank_by_code.setdefault(item["code"], {"items": [], "amount": Decimal("0.00")})
        group["items"].append(item)
        group["amount"] += item["amount"]

    used_bank_codes = set()
    for ticket_no, drug in drug_group.items():
        match_type = ""
        bank_code = ""
        bank = None
        if ticket_no in bank_by_code:
            bank_code = ticket_no
            bank = bank_by_code[bank_code]
            match_type = "完全匹配"
        else:
            unused_bank = {code: item for code, item in bank_by_code.items() if code not in used_bank_codes}
            fuzzy_code = find_fuzzy_match(ticket_no, unused_bank)
            if fuzzy_code:
                bank_code = fuzzy_code
                bank = bank_by_code[bank_code]
                match_type = "疑似匹配"
        supplier = "；".join(sorted(drug["suppliers"]))
        if bank:
            used_bank_codes.add(bank_code)
            bank_original = bank["amount"].quantize(Decimal("0.01"))
            bank_adjusted, rule = adjusted_bank_amount(bank_original, supplier, config)
            result, diff, note = compare_amounts(drug["amount"], bank_adjusted)
            if match_type == "疑似匹配":
                note = (note + "；" if note else "") + "票据号与银行编号非完全一致，请人工确认"
            for item in drug["items"]:
                item["used"] = True
            for item in bank["items"]:
                item["used"] = True
            results.append(make_reconcile_row("药店银行对账明细", result, match_type, ticket_no, bank_code, drug["amount"], bank_original, bank_adjusted, diff, note, drug["items"], bank["items"], supplier, rule))
        else:
            results.append(make_reconcile_row("药店孤立数据", "已上货/已收货，但银行未找到对应转账", "未匹配", ticket_no, "", drug["amount"], None, None, None, "银行交易用途中未找到对应货款编号", drug["items"], [], supplier))

    for code, bank in bank_by_code.items():
        if code in used_bank_codes:
            continue
        results.append(make_reconcile_row("银行孤立数据", "已转钱，但药店表未找到对应上货记录", "未匹配", "", code, None, bank["amount"], bank["amount"], None, "可能是票据号填错，或药店表未登记", [], bank["items"], ""))

    summary = {
        "药店票据号数量": len(drug_group),
        "银行带编号货款数量": len(bank_by_code),
        "完全匹配数量": sum(1 for item in results if item.match_type == "完全匹配"),
        "疑似匹配数量": sum(1 for item in results if item.match_type == "疑似匹配"),
        "药店孤立数量": sum(1 for item in results if item.category == "药店孤立数据"),
        "银行孤立数量": sum(1 for item in results if item.category == "银行孤立数据"),
        "高开票对账数量": sum(1 for item in results if item.category == "高开票差额对账"),
        "银行1笔付款公司对账数量": sum(1 for item in results if item.category == "银行1笔付款公司汇总"),
    }
    return results, summary, no_code_rows


def reconcile_drug_bank_final(
    drug_records: List[dict],
    bank_records: List[dict],
    high_invoice_records: List[dict],
    config: dict,
    current_branch_text: str = "",
) -> Tuple[List[ReconcileRow], dict, List[dict]]:
    """最终药店-银行对账：名单公司月汇总、海通固定系数、高开票按分部/日期/顺序配对。"""

    results: List[ReconcileRow] = []

    # 1. 银行1笔付款公司：不看货款编号，只按公司名称做月汇总。
    for company in config.get("monthly_companies", []):
        drug_items = group_drug_by_supplier(drug_records, company)
        bank_items = find_bank_by_company(bank_records, company)
        if not drug_items and not bank_items:
            continue

        drug_amount = summarize_records(drug_items)
        bank_original = summarize_records(bank_items)
        high_extra, high_items = take_monthly_high_invoices(company, bank_items, high_invoice_records, current_branch_text)
        bank_adjusted, rule = final_bank_amount(bank_original, company, config, high_extra, high_items)
        result, diff, note = compare_amounts(drug_amount, bank_adjusted)

        for item in drug_items:
            item["used"] = True
        for item in bank_items:
            item["used"] = True

        results.append(
            make_reconcile_row(
                "银行1笔付款公司汇总",
                result,
                "公司月汇总",
                "公司月汇总",
                "",
                drug_amount,
                bank_original,
                bank_adjusted,
                diff,
                note,
                drug_items,
                bank_items,
                company,
                rule,
            )
        )

    # 2. 海通/高开票/无编码付款：按公司、日期、顺序和金额匹配。
    supplier_names = sorted({item["supplier"] for item in drug_records if not item["used"] and item.get("supplier")})
    for supplier in supplier_names:
        drug_items_all = sorted(
            group_drug_by_supplier(drug_records, supplier),
            key=lambda item: (record_date(item, "date") or _dt.datetime.min, item["row"]),
        )
        bank_items_all = sorted(
            find_bank_by_company(bank_records, supplier),
            key=lambda item: (record_date(item, "time") or _dt.datetime.min, item["row"]),
        )
        high_items = high_invoice_candidates(high_invoice_records, supplier, current_branch_text)
        factor_company = is_special_multiplier_company(supplier, config)

        if not bank_items_all:
            continue
        if not (factor_company or high_items or any(not item.get("code") for item in bank_items_all)):
            continue

        if factor_company or high_items:
            drug_items = drug_items_all
            bank_items = bank_items_all
        else:
            drug_items = [item for item in drug_items_all if not item.get("ticket_no")]
            bank_items = [item for item in bank_items_all if not item.get("code")]

        if not drug_items and not bank_items:
            continue

        bank_entries = []
        for idx, bank in enumerate(bank_items):
            high_extra, matched_high = (
                (Decimal("0.00"), [])
                if factor_company
                else take_high_invoice_for_bank(bank, high_items, idx)
            )
            adjusted, rule = final_bank_amount(bank["amount"], supplier, config, high_extra, matched_high)
            bank_entries.append(
                {
                    "bank": bank,
                    "adjusted": adjusted,
                    "rule": rule,
                    "high_items": matched_high,
                }
            )

        unused_drugs = list(drug_items)
        for entry in bank_entries:
            bank = entry["bank"]
            if unused_drugs:
                bank_dt = record_date(bank, "time")
                best = min(
                    unused_drugs,
                    key=lambda drug: (
                        abs(drug["amount"] - entry["adjusted"]),
                        days_between(record_date(drug, "date"), bank_dt),
                        drug["row"],
                    ),
                )
                related_drugs = [best]
                unused_drugs.remove(best)
            else:
                related_drugs = []

            drug_amount = summarize_records(related_drugs) if related_drugs else Decimal("0.00")
            if related_drugs:
                result, diff, note = compare_amounts(drug_amount, entry["adjusted"])
            else:
                result, diff, note = "药店未找到", None, "按公司名称找到银行付款，但药店表中没有可配对的同供应商记录"

            bank["used"] = True
            for drug in related_drugs:
                drug["used"] = True

            match_type = "海通固定系数" if factor_company else "公司+日期/顺序"
            if entry["high_items"]:
                match_type += "+高开票"

            results.append(
                make_reconcile_row(
                    "药店银行对账明细",
                    result,
                    match_type,
                    "公司配对" if not related_drugs else related_drugs[0].get("ticket_no") or "公司配对",
                    bank.get("code", ""),
                    drug_amount if related_drugs else None,
                    bank["amount"],
                    entry["adjusted"],
                    diff,
                    note,
                    related_drugs,
                    [bank],
                    supplier,
                    entry["rule"],
                )
            )

        for drug in unused_drugs:
            results.append(
                make_reconcile_row(
                    "药店孤立数据",
                    "已上货/已收货，但银行未找到对应转账",
                    "公司配对未匹配",
                    drug.get("ticket_no") or "无票据号",
                    "",
                    drug["amount"],
                    None,
                    None,
                    None,
                    "同供应商银行付款数量少于药店记录，或付款日期/金额无法对应",
                    [drug],
                    [],
                    supplier,
                )
            )
            drug["used"] = True

    # 3. 普通票据号对账：剩余记录才按银行用途中的货款编号匹配。
    drug_group: Dict[str, dict] = {}
    for item in drug_records:
        if item["used"] or not item["ticket_no"]:
            continue
        group = drug_group.setdefault(item["ticket_no"], {"items": [], "amount": Decimal("0.00"), "suppliers": set()})
        group["items"].append(item)
        group["amount"] += item["amount"]
        if item["supplier"]:
            group["suppliers"].add(item["supplier"])

    bank_by_code: Dict[str, dict] = {}
    no_code_rows = []
    for item in bank_records:
        if item["used"]:
            continue
        if not item["code"]:
            no_code_rows.append(item)
            continue
        group = bank_by_code.setdefault(item["code"], {"items": [], "amount": Decimal("0.00")})
        group["items"].append(item)
        group["amount"] += item["amount"]

    used_bank_codes = set()
    for ticket_no, drug in drug_group.items():
        match_type = ""
        bank_code = ""
        bank = None
        if ticket_no in bank_by_code:
            bank_code = ticket_no
            bank = bank_by_code[bank_code]
            match_type = "完全匹配"
        else:
            unused_bank = {code: item for code, item in bank_by_code.items() if code not in used_bank_codes}
            fuzzy_code = find_fuzzy_match(ticket_no, unused_bank)
            if fuzzy_code:
                bank_code = fuzzy_code
                bank = bank_by_code[bank_code]
                match_type = "疑似匹配"

        supplier = "；".join(sorted(drug["suppliers"]))
        if bank:
            used_bank_codes.add(bank_code)
            bank_original = bank["amount"].quantize(Decimal("0.01"))
            bank_adjusted, rule = final_bank_amount(bank_original, supplier, config)
            result, diff, note = compare_amounts(drug["amount"], bank_adjusted)
            if match_type == "疑似匹配":
                note = (note + "；" if note else "") + "票据号与银行编号非完全一致，请人工确认"
            for item in drug["items"]:
                item["used"] = True
            for item in bank["items"]:
                item["used"] = True
            results.append(
                make_reconcile_row(
                    "药店银行对账明细",
                    result,
                    match_type,
                    ticket_no,
                    bank_code,
                    drug["amount"],
                    bank_original,
                    bank_adjusted,
                    diff,
                    note,
                    drug["items"],
                    bank["items"],
                    supplier,
                    rule,
                )
            )
        else:
            results.append(
                make_reconcile_row(
                    "药店孤立数据",
                    "已上货/已收货，但银行未找到对应转账",
                    "未匹配",
                    ticket_no,
                    "",
                    drug["amount"],
                    None,
                    None,
                    None,
                    "银行交易用途中未找到对应货款编号",
                    drug["items"],
                    [],
                    supplier,
                )
            )

    for code, bank in bank_by_code.items():
        if code in used_bank_codes:
            continue
        results.append(
            make_reconcile_row(
                "银行孤立数据",
                "已转钱，但药店表未找到对应上货记录",
                "未匹配",
                "",
                code,
                None,
                bank["amount"],
                bank["amount"],
                None,
                "可能是票据号填错，或药店表未登记",
                [],
                bank["items"],
                "",
            )
        )

    for item in no_code_rows:
        results.append(
            make_reconcile_row(
                "银行无编号数据",
                "银行已付款但未填写货款编号",
                "无编号",
                "",
                "",
                None,
                item["amount"],
                item["amount"],
                None,
                "不在银行1笔名单内，也未能按公司名称自动配对；请人工核对",
                [],
                [item],
                "",
            )
        )

    unmatched_high = [
        item
        for item in high_invoice_records
        if not item["used"] and item.get("amount", Decimal("0.00")) != Decimal("0.00")
    ]
    for item in unmatched_high:
        results.append(
            make_reconcile_row(
                "高开票未匹配数据",
                "高开票未参与自动对账",
                "分部/公司/日期未匹配",
                item.get("doc_no") or f"高开票行{item['row']}",
                "",
                None,
                None,
                item["amount"],
                None,
                f"供货商：{item.get('supplier','')}；分部：{item.get('branch','')}。未与本次对账分部或银行付款匹配。",
                [],
                [],
                item.get("supplier", ""),
                "未加到银行对账金额",
            )
        )

    summary = {
        "药店票据号数量": len(drug_group),
        "银行带编号货款数量": len(bank_by_code),
        "完全匹配数量": sum(1 for item in results if item.match_type == "完全匹配"),
        "疑似匹配数量": sum(1 for item in results if item.match_type == "疑似匹配"),
        "药店孤立数量": sum(1 for item in results if item.category == "药店孤立数据"),
        "银行孤立数量": sum(1 for item in results if item.category == "银行孤立数据"),
        "银行无编号数量": len(no_code_rows),
        "高开票参与对账数量": sum(1 for item in high_invoice_records if item["used"]),
        "高开票未匹配数量": len(unmatched_high),
        "银行1笔付款公司对账数量": sum(1 for item in results if item.category == "银行1笔付款公司汇总"),
    }
    return results, summary, no_code_rows


def issue_priority(item: ReconcileRow) -> int:
    """结果表排序用：越小越需要优先查看。"""

    if item.category == "药店孤立数据":
        return 1
    if item.category == "银行孤立数据":
        return 2
    if item.result in ("药店金额大于银行金额", "银行金额大于药店金额"):
        return 3
    if item.match_type == "疑似匹配":
        return 4
    return 9


def issue_reason(item: ReconcileRow) -> str:
    """把技术性的分类翻译成业务人员更容易看的问题原因。"""

    if item.category == "药店孤立数据":
        return "药店有票据号，银行未找到对应货款编号"
    if item.category == "银行孤立数据":
        return "银行已付款，药店表未找到对应票据号"
    if item.result == "药店金额大于银行金额":
        return "金额不一致：药店应结算大于银行支出"
    if item.result == "银行金额大于药店金额":
        return "金额不一致：银行支出大于药店应结算"
    if item.match_type == "疑似匹配":
        return "编号疑似一致，但不是完全相同"
    return "无异常"


def suggested_action(item: ReconcileRow) -> str:
    """给结果表增加可执行的处理建议。"""

    if item.category == "药店孤立数据":
        return "核对是否漏付款，或银行交易用途未填写正确票据号"
    if item.category == "银行孤立数据":
        return "核对是否已付款未上货，或银行交易用途编号填错"
    if item.result == "药店金额大于银行金额":
        return "核对是否少付款、分笔付款未填编号，或药店金额录入偏大"
    if item.result == "银行金额大于药店金额":
        return "核对是否多付款、药店漏录入，或银行付款金额填错"
    if item.match_type == "疑似匹配":
        return "人工确认票据号和银行编号是否属于同一笔业务"
    return "无需处理"


def reconcile_drug_bank(
    drug_data: Dict[str, dict], bank_data: Dict[str, dict]
) -> Tuple[List[ReconcileRow], dict]:
    """
    执行药店-银行对账。

    输出放在同一个 sheet 中，通过“数据类别”区分：
    - 匹配明细
    - 药店孤立数据
    - 银行孤立数据
    """

    results: List[ReconcileRow] = []
    used_bank_codes = set()

    # 先做完全匹配，再做疑似匹配。
    for ticket_no, drug in drug_data.items():
        match_type = ""
        bank_code = ""
        bank = None

        if ticket_no in bank_data:
            bank_code = ticket_no
            bank = bank_data[bank_code]
            match_type = "完全匹配"
        else:
            unused_bank = {code: item for code, item in bank_data.items() if code not in used_bank_codes}
            fuzzy_code = find_fuzzy_match(ticket_no, unused_bank)
            if fuzzy_code:
                bank_code = fuzzy_code
                bank = bank_data[bank_code]
                match_type = "疑似匹配"

        if bank:
            used_bank_codes.add(bank_code)
            result, diff, note = compare_amounts(drug["amount"], bank["amount"])
            if match_type == "疑似匹配":
                note = (note + "；" if note else "") + "票据号与银行编号非完全一致，请人工确认"

            results.append(
                ReconcileRow(
                    category="药店银行对账明细",
                    result=result,
                    match_type=match_type,
                    ticket_no=ticket_no,
                    bank_code=bank_code,
                    drug_amount=drug["amount"],
                    bank_amount=bank["amount"],
                    diff=diff,
                    note=note,
                    drug_rows=", ".join(drug["rows"]),
                    bank_rows=", ".join(bank["rows"]),
                    supplier="；".join(sorted(drug["suppliers"])),
                    bank_counterparty="；".join(sorted(bank["counterparties"])),
                    bank_purpose="；".join(bank["purposes"]),
                    bank_time="；".join(bank["times"]),
                )
            )
        else:
            results.append(
                ReconcileRow(
                    category="药店孤立数据",
                    result="已上货/已收货，但银行未找到对应转账",
                    match_type="未匹配",
                    ticket_no=ticket_no,
                    bank_code="",
                    drug_amount=drug["amount"],
                    bank_amount=None,
                    diff=None,
                    note="银行交易用途中未找到对应货款编号",
                    drug_rows=", ".join(drug["rows"]),
                    supplier="；".join(sorted(drug["suppliers"])),
                )
            )

    for bank_code, bank in bank_data.items():
        if bank_code in used_bank_codes:
            continue
        results.append(
            ReconcileRow(
                category="银行孤立数据",
                result="已转钱，但药店表未找到对应上货记录",
                match_type="未匹配",
                ticket_no="",
                bank_code=bank_code,
                drug_amount=None,
                bank_amount=bank["amount"],
                diff=None,
                note="可能是票据号填错，或药店表未登记",
                bank_rows=", ".join(bank["rows"]),
                bank_counterparty="；".join(sorted(bank["counterparties"])),
                bank_purpose="；".join(bank["purposes"]),
                bank_time="；".join(bank["times"]),
            )
        )

    summary = {
        "药店票据号数量": len(drug_data),
        "银行带编号货款数量": len(bank_data),
        "完全匹配数量": sum(1 for item in results if item.match_type == "完全匹配"),
        "疑似匹配数量": sum(1 for item in results if item.match_type == "疑似匹配"),
        "药店孤立数量": sum(1 for item in results if item.category == "药店孤立数据"),
        "银行孤立数量": sum(1 for item in results if item.category == "银行孤立数据"),
    }
    return results, summary


def make_medical_placeholder(bank_file: Path, declare_file: Path, config: Optional[dict] = None) -> List[List[object]]:
    """根据医保对账规则生成银行收入与申报表金额的对账结果。"""

    try:
        bank_rows = load_medical_bank_income_rows(bank_file)
        current_declare_values, current_period = load_declare_medical_values(declare_file)
        ensure_declare_in_history(declare_file, current_period)
        declare_map = load_history_declare_map(declare_file, current_declare_values, current_period)
        rows = build_medical_reconcile_rows(bank_rows, declare_map, current_period, config or {})
        if not rows:
            return [["", "", "", "", "", "", None, "", None, None, None, "银行未找到", "未找到可用于医保对账的银行收入记录"]]
        return rows
    except Exception as exc:
        return [["", "", "", "", "", "", None, "", None, None, None, "医保对账失败", f"{exc}"]]


def strip_xml_tags(text: str) -> str:
    """去掉银行交易用途中的简单 XML 标签，便于关键字匹配。"""

    return re.sub(r"<[^>]+>", "", normalize_text(text)).strip(" .。")


def extract_declare_period(workbook) -> Tuple[Optional[int], Optional[int], str, str]:
    """
    从申报表标题行里识别结算年月。

    返回：年、月、YYYYMM、中文月份文本。例如 2026、4、202604、4月。
    """

    for ws in workbook.worksheets:
        for row in range(1, min(ws.max_row, 6) + 1):
            text = " ".join(normalize_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 8) + 1))
            match = re.search(r"(\d{4})年\s*(\d{1,2})月", text)
            if match:
                year = int(match.group(1))
                month = int(match.group(2))
                return year, month, f"{year}{month:02d}", f"{month}月"
    return None, None, "", ""


def extract_declare_unit(workbook) -> str:
    """从申报表中识别“报送单位”。"""

    for ws in workbook.worksheets:
        for row in range(1, min(ws.max_row, 6) + 1):
            text = " ".join(normalize_text(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 10) + 1))
            match = re.search(r"报送单位[:：]\s*(.+?)(?:\s+结算时段|$)", text)
            if match:
                return normalize_text(match.group(1))
    return ""


def extract_bank_account_name(bank_file: Path) -> str:
    """从银行账户明细表头识别户名。"""

    df = read_excel_any(bank_file, header=None)
    for row_index in range(min(len(df), 5)):
        row_text = " ".join(normalize_text(v) for v in df.iloc[row_index].tolist())
        match = re.search(r"户名[:：]\s*(.+?)(?:\s+币种[:：]|\s*$)", row_text)
        if match:
            return normalize_text(match.group(1))
    return ""


def cell_decimal(ws, cell_ref: str, label: str) -> Decimal:
    """读取申报表固定单元格金额。"""

    return to_decimal(ws[cell_ref].value, label)


def pick_sheet(workbook, keyword: str, fallback_index: int):
    """按工作表名称关键字选择 sheet，找不到时用固定顺序兜底。"""

    for ws in workbook.worksheets:
        if keyword in ws.title:
            return ws
    return workbook.worksheets[fallback_index]


def load_declare_medical_values(declare_file: Path) -> Tuple[Dict[str, Decimal], dict]:
    """
    读取申报表中医保对账所需的金额项。

    当前申报表模板是固定结构，所以这里按模板单元格取值；如果以后模板行列变化，
    只需要修改这个函数里的单元格映射。
    """

    wb = load_workbook(declare_file, data_only=True, read_only=False)
    year, month, ym, month_text = extract_declare_period(wb)
    unit = extract_declare_unit(wb)
    ws_local = pick_sheet(wb, "居民、职工", 0)
    ws_remote = pick_sheet(wb, "异地", 1 if len(wb.worksheets) > 1 else 0)

    values = {
        "定点药店购药-个人账户-职工": cell_decimal(ws_local, "H6", "定点药店购药-个人账户-职工"),
        "普通门诊-个人账户-职工": cell_decimal(ws_local, "Q6", "普通门诊-个人账户-职工"),
        "门诊慢特病-个人账户-职工": cell_decimal(ws_local, "Z6", "门诊慢特病-个人账户-职工"),
        "门诊慢特病-医疗救助-居民": cell_decimal(ws_local, "AC5", "门诊慢特病-医疗救助-居民"),
        "门诊慢特病-医疗救助-职工": cell_decimal(ws_local, "AC6", "门诊慢特病-医疗救助-职工"),
        "门诊慢特病-基本医保-居民": cell_decimal(ws_local, "Y5", "门诊慢特病-基本医保-居民"),
        "门诊慢特病-基本医保-职工": cell_decimal(ws_local, "Y6", "门诊慢特病-基本医保-职工"),
        "普通门诊-基本医保-职工": cell_decimal(ws_local, "P6", "普通门诊-基本医保-职工"),
        "定点药店购药-家庭共济-居民": cell_decimal(ws_local, "I5", "定点药店购药-家庭共济-居民"),
        "门诊慢特病-家庭共济-居民": cell_decimal(ws_local, "AA5", "门诊慢特病-家庭共济-居民"),
        "定点药店购药-家庭共济-职工": cell_decimal(ws_local, "I6", "定点药店购药-家庭共济-职工"),
        "普通门诊-家庭共济-职工": cell_decimal(ws_local, "R6", "普通门诊-家庭共济-职工"),
        "门诊慢特病-家庭共济-职工": cell_decimal(ws_local, "AA6", "门诊慢特病-家庭共济-职工"),
        "省内异地-统筹基金-居民": cell_decimal(ws_remote, "G5", "省内异地-统筹基金-居民"),
        "省内异地-统筹基金-职工": cell_decimal(ws_remote, "G6", "省内异地-统筹基金-职工"),
        "省内异地-个人账户-职工": cell_decimal(ws_remote, "H6", "省内异地-个人账户-职工"),
        "跨省异地-个人账户-职工": cell_decimal(ws_remote, "O6", "跨省异地-个人账户-职工"),
        "跨省异地-统筹基金-职工": cell_decimal(ws_remote, "N6", "跨省异地-统筹基金-职工"),
        "跨省异地-统筹基金-居民": cell_decimal(ws_remote, "N5", "跨省异地-统筹基金-居民"),
        "跨省异地-家庭共济-居民": cell_decimal(ws_remote, "P5", "跨省异地-家庭共济-居民"),
    }
    return values, {"year": year, "month": month, "ym": ym, "month_text": month_text, "unit": unit, "file": str(declare_file)}


def get_history_dir() -> Path:
    """历史申报表库目录，默认放在 EXE/源码同级。"""

    return get_base_dir() / HISTORY_FOLDER_NAME


def ensure_declare_in_history(declare_file: Path, period: dict) -> None:
    """
    把当前申报表保存到历史申报表库。

    这样后续银行流水出现历史月份款项时，工具可以自动找到对应月份申报表。
    """

    ym = period.get("ym")
    unit = period.get("unit") or "未知单位"
    if not ym or not declare_file.exists():
        return
    history_dir = get_history_dir()
    month_dir = history_dir / ym / safe_path_name(unit)
    month_dir.mkdir(parents=True, exist_ok=True)
    target = month_dir / "申报表.xlsx"
    try:
        if not target.exists() or declare_file.stat().st_mtime > target.stat().st_mtime:
            import shutil

            shutil.copy2(declare_file, target)
    except PermissionError:
        # 历史库写入失败不影响本次对账，只是后续跨月匹配可能缺少资料。
        return


def import_declare_to_history(declare_file: Path) -> Tuple[str, Path]:
    """手动导入申报表到历史库，供界面按钮调用。"""

    values, period = load_declare_medical_values(declare_file)
    ensure_declare_in_history(declare_file, period)
    ym = period.get("ym") or "未知月份"
    unit = period.get("unit") or "未知单位"
    return ym, get_history_dir() / ym / safe_path_name(unit) / "申报表.xlsx"


def load_history_declare_map(
    current_declare_file: Path, current_values: Dict[str, Decimal], current_period: dict
) -> Dict[str, dict]:
    """
    读取历史申报表库，返回 {YYYYMM: {values, period}}。

    当前选择的申报表一定加入 map，即便历史库写入失败，本次也可正常对账。
    """

    result: Dict[str, dict] = {}
    current_ym = current_period.get("ym")
    current_unit = current_period.get("unit") or ""
    if current_ym:
        result[current_ym] = {"values": current_values, "period": current_period}

    history_dir = get_history_dir()
    if not history_dir.exists():
        return result

    for file_path in history_dir.glob("*/*/申报表.xlsx"):
        try:
            values, period = load_declare_medical_values(file_path)
        except Exception:
            continue
        ym = period.get("ym")
        unit = period.get("unit") or ""
        if ym and (not current_unit or unit == current_unit):
            result[ym] = {"values": values, "period": period}
    return result


def load_medical_bank_income_rows(bank_file: Path) -> List[dict]:
    """读取银行收入记录，供医保对账规则匹配。"""

    df = read_excel_any(bank_file, header=2)
    df.columns = [normalize_text(col) for col in df.columns]
    require_columns(df, ["交易时间", "收入金额", "对方户名", "交易用途"], "银行数据表")

    rows: List[dict] = []
    for index, row in df.iterrows():
        purpose = strip_xml_tags(row.get("交易用途"))
        income_raw = row.get("收入金额")
        # 银行表底部可能有“总收入金额”等汇总文字行，跳过即可。
        try:
            income = to_decimal(income_raw, "银行收入金额")
        except UserVisibleError:
            if not purpose:
                continue
            raise
        if income == Decimal("0.00"):
            continue
        if not purpose or "总收入" in purpose or "往来款" in purpose:
            continue
        rows.append(
            {
                "row": index + 4,
                "time": normalize_text(row.get("交易时间")),
                "amount": income,
                "counterparty": normalize_text(row.get("对方户名")),
                "purpose": purpose,
                "business_ym": "",
            }
        )
    return rows


def infer_bank_business_ym(purpose: str, current_period: dict) -> str:
    """从银行备注中识别业务月份。"""

    text = normalize_text(purpose)
    match = re.search(r"(20\d{2})(0[1-9]|1[0-2])", text)
    if match:
        return match.group(1) + match.group(2)

    month_match = re.search(r"(\d{1,2})月", text)
    if month_match and current_period.get("year"):
        month = int(month_match.group(1))
        if 1 <= month <= 12:
            return f"{int(current_period['year'])}{month:02d}"

    return current_period.get("ym", "")


def amount_sum(values: Dict[str, Decimal], items: Iterable[str]) -> Decimal:
    """多个申报表金额项求和。"""

    total = Decimal("0.00")
    for item in items:
        total += values.get(item, Decimal("0.00"))
    return total.quantize(Decimal("0.01"))


def medical_status(bank_amount: Optional[Decimal], declare_amount: Decimal) -> Tuple[str, Optional[Decimal], str]:
    """医保金额比较。"""

    if bank_amount is None:
        return "银行未找到", None, "银行收入中未找到对应备注"
    diff = (bank_amount - declare_amount).quantize(Decimal("0.01"))
    if diff == Decimal("0.00"):
        return "对账成功", diff, ""
    if abs(diff) <= TINY_DIFF_THRESHOLD:
        return "金额极小差异", diff, "差异很小，请人工确认"
    return "金额不一致", diff, "银行收入与申报表金额不一致，请核对"


def filter_bank_rows(bank_rows: List[dict], period: dict, matcher: Callable[[str], bool], need_month: bool = True) -> List[dict]:
    """按备注关键字和申报月份筛选银行收入记录。"""

    ym = period.get("ym", "")
    month_text = period.get("month_text", "")
    matched = []
    for row in bank_rows:
        purpose = row["purpose"]
        if need_month and ym and month_text and (ym not in purpose and month_text not in purpose):
            continue
        if matcher(purpose):
            matched.append(row)
    return matched


def filter_bank_rows_by_ym(bank_rows: List[dict], ym: str, matcher: Callable[[str], bool]) -> List[dict]:
    """按业务月份和备注规则筛选银行医保收入。"""

    return [row for row in bank_rows if row.get("business_ym") == ym and matcher(row["purpose"])]


def apply_medical_discount(target_name: str, amount: Decimal, config: dict) -> Tuple[Decimal, Decimal, str]:
    """医保申报金额按配置折扣后用于对账。返回原始金额、对账金额、规则。"""

    factor = Decimal("1.00")
    for item, configured_factor in config.get("medical_discount_items", {}).items():
        if item and item in target_name:
            factor = configured_factor
            break
    adjusted = (amount * factor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    rule = f"申报金额×{factor}" if factor != Decimal("1.00") else ""
    return amount, adjusted, rule


def append_medical_row(
    rows: List[List[object]],
    business_ym: str,
    rule_name: str,
    bank_items: List[dict],
    bank_amount: Optional[Decimal],
    target_name: str,
    declare_amount: Decimal,
    note: str = "",
    config: Optional[dict] = None,
) -> None:
    """追加一行医保对账结果。"""

    config = config or {}
    declare_original, declare_check_amount, discount_rule = apply_medical_discount(target_name, declare_amount, config)
    status, diff, auto_note = medical_status(bank_amount, declare_check_amount)
    row_numbers = "；".join(str(item["row"]) for item in bank_items)
    purposes = "；".join(item["purpose"] for item in bank_items)
    times = "；".join(item["time"] for item in bank_items)
    counterparties = "；".join(sorted({item["counterparty"] for item in bank_items if item["counterparty"]}))
    rows.append(
        [
            business_ym,
            rule_name,
            row_numbers,
            times,
            counterparties,
            purposes,
            format_money(bank_amount),
            target_name,
            format_money(declare_original),
            format_money(declare_check_amount),
            format_money(diff),
            status,
            "；".join(x for x in [discount_rule, note or auto_note] if x),
        ]
    )


def append_sum_rule(
    rows: List[List[object]],
    business_ym: str,
    rule_name: str,
    bank_items: List[dict],
    target_name: str,
    declare_amount: Decimal,
    config: Optional[dict] = None,
) -> None:
    """银行多笔求和后与申报表目标金额对账。"""

    bank_amount = sum((item["amount"] for item in bank_items), Decimal("0.00")).quantize(Decimal("0.01")) if bank_items else None
    append_medical_row(rows, business_ym, rule_name, bank_items, bank_amount, target_name, declare_amount, "银行侧按规则求和后对账", config)


def append_best_match_rule(
    rows: List[List[object]],
    business_ym: str,
    rule_name: str,
    bank_items: List[dict],
    targets: List[Tuple[str, Decimal]],
    config: Optional[dict] = None,
) -> None:
    """
    多笔银行记录与多个申报项目做最优匹配。

    优先完全一致；没有完全一致时，选择差异最小的一组。
    """

    unused = list(bank_items)
    for target_name, declare_amount in targets:
        if not unused:
            append_medical_row(rows, business_ym, rule_name, [], None, target_name, declare_amount, "申报表有金额，但银行未找到对应打款", config)
            continue

        _, check_amount, _ = apply_medical_discount(target_name, declare_amount, config or {})
        exact_index = next((i for i, item in enumerate(unused) if item["amount"] == check_amount), None)
        if exact_index is None:
            exact_index = min(range(len(unused)), key=lambda i: abs(unused[i]["amount"] - check_amount))
            note = "未找到完全一致金额，已按差异最小原则匹配"
        else:
            note = "金额完全一致匹配"

        item = unused.pop(exact_index)
        append_medical_row(rows, business_ym, rule_name, [item], item["amount"], target_name, declare_amount, note, config)

    for item in unused:
        append_medical_row(rows, business_ym, rule_name, [item], item["amount"], "未配置对应申报项目", Decimal("0.00"), "银行侧多出一笔同类收入", config)


def build_medical_reconcile_rows(bank_rows: List[dict], declare_map: Dict[str, dict], current_period: dict, config: dict) -> List[List[object]]:
    """以申报表项目为主体，逐月生成医保对账明细。"""

    rows: List[List[object]] = []
    used_bank_rows = set()

    for row in bank_rows:
        row["business_ym"] = infer_bank_business_ym(row["purpose"], current_period)

    for ym in sorted(declare_map):
        values = declare_map[ym]["values"]

        personal = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "两定结算个人账户" in p)
        append_best_match_rule(
            rows,
            ym,
            "两定结算个人账户",
            personal,
            [
                ("定点药店购药-个人账户-职工", amount_sum(values, ["定点药店购药-个人账户-职工"])),
                (
                    "普通门诊-个人账户-职工 + 门诊慢特病-个人账户-职工",
                    amount_sum(values, ["普通门诊-个人账户-职工", "门诊慢特病-个人账户-职工"]),
                ),
            ],
            config,
        )
        used_bank_rows.update(item["row"] for item in personal)

        medical_aid = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "两定结算医疗救助" in p)
        append_sum_rule(
            rows,
            ym,
            "两定结算医疗救助",
            medical_aid,
            "门诊慢特病-医疗救助-居民 + 门诊慢特病-医疗救助-职工",
            amount_sum(values, ["门诊慢特病-医疗救助-居民", "门诊慢特病-医疗救助-职工"]),
            config,
        )
        used_bank_rows.update(item["row"] for item in medical_aid)

        resident_fund = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "两定结算居民统筹" in p)
        append_sum_rule(rows, ym, "两定结算居民统筹", resident_fund, "门诊慢特病-基本医保-居民", amount_sum(values, ["门诊慢特病-基本医保-居民"]), config)
        used_bank_rows.update(item["row"] for item in resident_fund)

        staff_fund = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "两定结算职工统筹" in p)
        append_best_match_rule(
            rows,
            ym,
            "两定结算职工统筹",
            staff_fund,
            [
                ("门诊慢特病-基本医保-职工", amount_sum(values, ["门诊慢特病-基本医保-职工"])),
                ("普通门诊-基本医保-职工", amount_sum(values, ["普通门诊-基本医保-职工"])),
            ],
            config,
        )
        used_bank_rows.update(item["row"] for item in staff_fund)

        family = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "两定结算家庭共济" in p)
        append_best_match_rule(
            rows,
            ym,
            "两定结算家庭共济",
            family,
            [
                (
                    "定点药店购药-家庭共济-居民 + 门诊慢特病-家庭共济-居民",
                    amount_sum(values, ["定点药店购药-家庭共济-居民", "门诊慢特病-家庭共济-居民"]),
                ),
                ("定点药店购药-家庭共济-职工", amount_sum(values, ["定点药店购药-家庭共济-职工"])),
                (
                    "普通门诊-家庭共济-职工 + 门诊慢特病-家庭共济-职工",
                    amount_sum(values, ["普通门诊-家庭共济-职工", "门诊慢特病-家庭共济-职工"]),
                ),
            ],
            config,
        )
        used_bank_rows.update(item["row"] for item in family)

        province_slow = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "省内" in p and ("慢病" in p or "慢特" in p))
        append_sum_rule(rows, ym, "省内慢病/慢特", province_slow, "省内异地-统筹基金-居民", amount_sum(values, ["省内异地-统筹基金-居民"]), config)
        used_bank_rows.update(item["row"] for item in province_slow)

        province_normal_slow = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "省内" in p and ("普通门诊" in p or "慢特" in p))
        append_sum_rule(
            rows,
            ym,
            "省内普通门诊 + 省内慢特",
            province_normal_slow,
            "省内异地-统筹基金-职工 + 省内异地-个人账户-职工",
            amount_sum(values, ["省内异地-统筹基金-职工", "省内异地-个人账户-职工"]),
            config,
        )
        used_bank_rows.update(item["row"] for item in province_normal_slow)

        cross_staff = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "跨省异地" in p and "居民" not in p)
        append_sum_rule(
            rows,
            ym,
            "跨省异地",
            cross_staff,
            "跨省异地-个人账户-职工 + 跨省异地-统筹基金-职工",
            amount_sum(values, ["跨省异地-个人账户-职工", "跨省异地-统筹基金-职工"]),
            config,
        )
        used_bank_rows.update(item["row"] for item in cross_staff)

        cross_resident = filter_bank_rows_by_ym(bank_rows, ym, lambda p: "跨省异地居民" in p)
        append_sum_rule(
            rows,
            ym,
            "跨省异地居民",
            cross_resident,
            "跨省异地-统筹基金-居民 + 跨省异地-家庭共济-居民",
            amount_sum(values, ["跨省异地-统筹基金-居民", "跨省异地-家庭共济-居民"]),
            config,
        )
        used_bank_rows.update(item["row"] for item in cross_resident)

    for item in bank_rows:
        ym = item.get("business_ym", "")
        if item["row"] in used_bank_rows or not ym or ym in declare_map:
            continue
        current_unit = current_period.get("unit") or "当前报送单位"
        append_medical_row(
            rows,
            ym,
            "历史月份银行收入",
            [item],
            item["amount"],
            f"未找到 {ym} 申报表",
            Decimal("0.00"),
            f"银行出现 {ym} 的医保收入，请先导入 {current_unit} 的 {ym} 申报表到历史库",
            config,
        )

    return rows


def format_money(value: Optional[Decimal]) -> Optional[float]:
    """给 openpyxl 写入金额用。"""

    if value is None:
        return None
    return float(value)


def add_table(ws, table_name: str) -> None:
    """给工作表数据区域添加 Excel 表格样式和筛选。"""

    if ws.max_row < 2 or ws.max_column < 1:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


def style_sheet(ws, money_columns: Iterable[int] = ()) -> None:
    """统一美化工作表：标题、冻结、列宽、边框、筛选、颜色。"""

    header_fill = PatternFill("solid", fgColor="244062")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    for col_idx in money_columns:
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for item in cell:
                item.number_format = '#,##0.00'

    # 根据“问题原因/结果”做颜色提示，方便第一眼找到需要处理的数据。
    result_col = None
    reason_col = None
    for cell in ws[1]:
        if cell.value == "对账结果":
            result_col = cell.column
        if cell.value == "判断结果":
            result_col = cell.column
        if cell.value == "问题原因":
            reason_col = cell.column
    color_source_col = reason_col or result_col
    if color_source_col:
        fills = {
            "无异常": PatternFill("solid", fgColor="E2F0D9"),
            "药店有票据号，银行未找到对应货款编号": PatternFill("solid", fgColor="FCEEEF"),
            "银行已付款，药店表未找到对应票据号": PatternFill("solid", fgColor="FCE4D6"),
            "金额不一致：药店应结算大于银行支出": PatternFill("solid", fgColor="FFF2CC"),
            "金额不一致：银行支出大于药店应结算": PatternFill("solid", fgColor="FCE4D6"),
            "编号疑似一致，但不是完全相同": PatternFill("solid", fgColor="D9EAD3"),
            "对账成功": PatternFill("solid", fgColor="E2F0D9"),
            "金额不一致": PatternFill("solid", fgColor="FFF2CC"),
            "金额极小差异": PatternFill("solid", fgColor="EADCF8"),
            "银行未找到": PatternFill("solid", fgColor="FCEEEF"),
        }
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, color_source_col).value
            fill = fills.get(value)
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row, col).fill = fill

    # 前几列是人工处理入口，居中显示，读起来更稳。
    for col in range(1, min(ws.max_column, 5) + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_cells in ws.columns:
        max_len = 8
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(text), 45))
        ws.column_dimensions[col_letter].width = max_len + 2


def export_results(
    output_dir: Path,
    reconcile_rows: List[ReconcileRow],
    medical_rows: List[List[object]],
    summary: dict,
    no_code_rows: List[dict],
    source_files: FileSelection,
) -> Path:
    """导出最终结果 Excel。"""

    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = _dt.datetime.now().strftime("%Y%m%d_%H%M")
    output_path = output_dir / f"{OUTPUT_PREFIX}_{timestamp}.xlsx"

    wb = Workbook()

    ws = wb.active
    ws.title = "药店银行对账"
    display_rows = sorted(reconcile_rows, key=lambda item: (issue_priority(item), item.ticket_no or item.bank_code))
    headers = [
        "处理优先级",
        "问题原因",
        "建议处理",
        "对账结果",
        "匹配方式",
        "药店票据号码",
        "银行货款编号",
        "药店应结算",
        "银行原始支出金额",
        "银行对账金额",
        "差额(药店-银行对账)",
        "金额调整规则",
        "备注",
        "药店原始行",
        "银行原始行",
        "药店供货商",
        "银行对方户名",
        "银行交易时间",
        "银行交易用途",
    ]
    ws.append(headers)
    for item in display_rows:
        ws.append(
            [
                issue_priority(item),
                issue_reason(item),
                suggested_action(item),
                item.result,
                item.match_type,
                item.ticket_no or "-",
                item.bank_code,
                format_money(item.drug_amount),
                format_money(item.bank_original_amount if item.bank_original_amount is not None else item.bank_amount),
                format_money(item.bank_adjusted_amount if item.bank_adjusted_amount is not None else item.bank_amount),
                format_money(item.diff),
                item.adjustment_rule,
                item.note,
                item.drug_rows,
                item.bank_rows,
                item.supplier,
                item.bank_counterparty,
                item.bank_time,
                item.bank_purpose,
            ]
        )
    add_table(ws, "DrugBankReconcile")
    style_sheet(ws, money_columns=[8, 9, 10, 11])
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["F"].width = 26
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["K"].width = 24
    ws.column_dimensions["L"].width = 30
    ws.column_dimensions["P"].width = 42
    ws.column_dimensions["R"].width = 55

    ws_med = wb.create_sheet("医保收款对账")
    ws_med.append([
        "业务月份",
        "规则名称",
        "银行原始行",
        "交易时间",
        "对方户名",
        "交易用途",
        "银行收入金额",
        "申报表对应项",
        "申报表原始金额",
        "申报表对账金额",
        "差额(银行-申报)",
        "判断结果",
        "备注",
    ])
    for row in medical_rows:
        ws_med.append(row)
    add_table(ws_med, "MedicalPlaceholder")
    style_sheet(ws_med, money_columns=[7, 9, 10, 11])
    ws_med.column_dimensions["A"].width = 22
    ws_med.column_dimensions["B"].width = 22
    ws_med.column_dimensions["F"].width = 48
    ws_med.column_dimensions["H"].width = 46

    ws_sum = wb.create_sheet("运行摘要")
    ws_sum.append(["项目", "内容"])
    ws_sum.append(["生成时间", _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_sum.append(["药店数据文件", str(source_files.drug_file or "")])
    ws_sum.append(["银行数据文件", str(source_files.bank_file or "")])
    ws_sum.append(["申报表文件", str(source_files.declare_file or "")])
    ws_sum.append(["银行户名", source_files.bank_account_name])
    ws_sum.append(["申报表报送单位", source_files.declare_unit])
    ws_sum.append(["单位校验结果", source_files.unit_check_message])
    for key, value in summary.items():
        ws_sum.append([key, value])
    ws_sum.append(["银行货款无编号数量", len(no_code_rows)])
    ws_sum.append(["说明", "银行交易用途只有“货款.”但没有编号的记录，不参与匹配。"])
    ws_sum.append(["医保收款说明", "已按医保对账规则配置表_修订版的规则内置对账。"])
    ws_sum.append(["AI接口预留", "后续可在源码 ai_assisted_mapping_placeholder 函数中接入 AI 辅助规则识别。"])
    style_sheet(ws_sum)

    wb.save(output_path)
    return output_path


def ai_assisted_mapping_placeholder(payload: dict) -> dict:
    """
    AI 接口预留位置。

    后续如果要接 AI，可以在这里：
    1. 把银行交易用途、申报表字段、历史人工映射规则组成 payload；
    2. 调用外部 AI 服务；
    3. 返回建议映射关系；
    4. 再由人工确认后写入医保对账配置。

    当前为了保证工具离线可用，不实际调用网络接口。
    """

    return {"enabled": False, "message": "AI 接口暂未启用", "payload": payload}


def run_reconcile(selection: FileSelection, log: Callable[[str], None]) -> Path:
    """一键对账主流程。"""

    log("开始读取数据文件...")
    drug_file = selection.drug_file
    bank_file = selection.bank_file
    declare_file = selection.declare_file

    if not drug_file or not bank_file or not declare_file:
        raise UserVisibleError("请先加载默认文件夹，或手动选择药店、银行、申报表三个文件。")

    config = load_business_config(selection.config_file)
    log("已加载业务规则配置。")

    log(f"药店数据：{drug_file.name}")
    drug_records = load_drug_records(drug_file)
    log(f"读取药店明细 {len(drug_records)} 行。")

    log(f"银行数据：{bank_file.name}")
    bank_account_name = extract_bank_account_name(bank_file)
    if bank_account_name:
        log(f"银行户名：{bank_account_name}")
    bank_records = load_bank_records(bank_file)
    log(f"读取银行支出明细 {len(bank_records)} 行。")

    high_invoice_records = load_high_invoice_records(selection.high_invoice_file)
    if high_invoice_records:
        log(f"读取高开票差额记录 {len(high_invoice_records)} 行。")

    log("开始执行药店-银行对账...")
    reconcile_rows, summary, no_code_rows = reconcile_drug_bank_final(
        drug_records, bank_records, high_invoice_records, config, bank_account_name
    )
    log(f"完全匹配 {summary['完全匹配数量']} 条，疑似匹配 {summary['疑似匹配数量']} 条。")
    log(f"药店孤立 {summary['药店孤立数量']} 条，银行孤立 {summary['银行孤立数量']} 条。")

    _, declare_period = load_declare_medical_values(declare_file)
    declare_unit = declare_period.get("unit", "")
    selection.declare_unit = declare_unit
    selection.bank_account_name = bank_account_name
    if declare_unit:
        log(f"申报表报送单位：{declare_unit}")
    if bank_account_name and declare_unit and bank_account_name != declare_unit:
        selection.unit_check_message = "银行户名与申报表报送单位不一致，请确认是否选错分部文件。"
        log("提示：" + selection.unit_check_message)
    else:
        selection.unit_check_message = "银行户名与申报表报送单位一致。" if bank_account_name and declare_unit else "未能完整识别银行户名或申报单位。"

    log("开始执行医保收款对账...")
    medical_rows = make_medical_placeholder(bank_file, declare_file, config)

    output_dir = selection.output_folder or get_base_dir()
    log("正在导出 Excel 结果...")
    output_path = export_results(output_dir, reconcile_rows, medical_rows, summary, no_code_rows, selection)
    save_settings(selection)
    log(f"处理完成：{output_path.name}")
    return output_path


class ReconcileApp(tk.Tk if tk else object):
    """桌面窗口。"""

    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1080x760")
        self.minsize(960, 680)
        self.selection = FileSelection()
        self.last_output_path: Optional[Path] = None
        load_settings_into_selection(self.selection)
        self._build_style()
        self._build_ui()
        self.refresh_vars_from_selection()
        if not (self.selection.drug_file or self.selection.bank_file or self.selection.declare_file):
            self.reload_default_folder()
        else:
            self.log("已恢复上次选择的文件路径，可直接核对后执行对账。")

    def _build_style(self) -> None:
        self.configure(bg="#EEF2F7")
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#EEF2F7")
        style.configure("Card.TFrame", background="#FFFFFF", relief="flat")
        style.configure("Toolbar.TFrame", background="#FFFFFF")
        style.configure("TLabel", background="#EEF2F7", foreground="#1F2937", font=("Microsoft YaHei UI", 10))
        style.configure("Card.TLabel", background="#FFFFFF", foreground="#1F2937", font=("Microsoft YaHei UI", 10))
        style.configure("Title.TLabel", background="#EEF2F7", foreground="#0F2E4F", font=("Microsoft YaHei UI", 20, "bold"))
        style.configure("Sub.TLabel", background="#EEF2F7", foreground="#64748B", font=("Microsoft YaHei UI", 10))
        style.configure("Section.TLabel", background="#FFFFFF", foreground="#0F2E4F", font=("Microsoft YaHei UI", 12, "bold"))
        style.configure("Hint.TLabel", background="#FFFFFF", foreground="#64748B", font=("Microsoft YaHei UI", 9))
        style.configure("TEntry", padding=(8, 6))
        style.configure("TButton", font=("Microsoft YaHei UI", 10), padding=(10, 7))
        style.configure("Primary.TButton", font=("Microsoft YaHei UI", 11, "bold"), padding=(18, 11), foreground="#FFFFFF", background="#1D4ED8")
        style.map("Primary.TButton", background=[("active", "#1E40AF"), ("pressed", "#1E3A8A")], foreground=[("disabled", "#E5E7EB")])
        style.configure("Quiet.TButton", font=("Microsoft YaHei UI", 10), padding=(10, 7))
        style.configure("Danger.TButton", font=("Microsoft YaHei UI", 10), padding=(10, 7), foreground="#991B1B")
        style.configure("TNotebook", background="#EEF2F7", borderwidth=0)
        style.configure("TNotebook.Tab", font=("Microsoft YaHei UI", 10), padding=(18, 8))

    def _build_ui(self) -> None:
        outer = ttk.Frame(self, padding=18)
        outer.pack(fill="both", expand=True)

        ttk.Label(outer, text=APP_TITLE, style="Title.TLabel").pack(anchor="w")
        ttk.Label(
            outer,
            text="选择本次数据、维护规则模板、生成带摘要和筛选的 Excel 对账结果。",
            style="Sub.TLabel",
        ).pack(anchor="w", pady=(4, 12))

        status = ttk.Frame(outer, style="Card.TFrame", padding=(16, 12))
        status.pack(fill="x", pady=(0, 12))
        ttk.Label(status, text="当前规则", style="Section.TLabel").pack(side="left")
        ttk.Label(
            status,
            text="银行1笔按公司月汇总；海通固定 ×1.22；高开票按供货商、分部、日期/顺序匹配。",
            style="Card.TLabel",
        ).pack(side="left", padx=(16, 0))

        self.notebook = ttk.Notebook(outer)
        self.notebook.pack(fill="both", expand=True)

        main_page = ttk.Frame(self.notebook, padding=14)
        config_page = ttk.Frame(self.notebook, padding=14)
        self.notebook.add(main_page, text="一键对账")
        self.notebook.add(config_page, text="规则配置")

        self.drug_var = tk.StringVar()
        self.bank_var = tk.StringVar()
        self.declare_var = tk.StringVar()
        self.output_var = tk.StringVar()
        self.config_var = tk.StringVar()
        self.high_invoice_var = tk.StringVar()

        file_card = ttk.Frame(main_page, style="Card.TFrame", padding=16)
        file_card.pack(fill="x", pady=(0, 12))
        ttk.Label(file_card, text="本次对账数据", style="Section.TLabel").grid(row=0, column=0, sticky="w", columnspan=3, pady=(0, 8))
        self._add_file_row(file_card, 1, "药店数据表", self.drug_var, lambda: self.choose_file("drug"))
        self._add_file_row(file_card, 2, "银行数据表", self.bank_var, lambda: self.choose_file("bank"))
        self._add_file_row(file_card, 3, "申报表", self.declare_var, lambda: self.choose_file("declare"))
        self._add_file_row(file_card, 4, "输出文件夹", self.output_var, self.choose_output_folder)

        action_card = ttk.Frame(main_page, style="Card.TFrame", padding=16)
        action_card.pack(fill="x", pady=(0, 12))
        ttk.Label(action_card, text="常用操作", style="Section.TLabel").pack(anchor="w", pady=(0, 10))
        row1 = ttk.Frame(action_card, style="Toolbar.TFrame")
        row1.pack(fill="x")
        ttk.Button(row1, text="一键执行对账", style="Primary.TButton", command=self.execute).pack(side="left")
        ttk.Button(row1, text="重新加载默认Excel", command=self.reload_default_folder).pack(side="left", padx=(10, 0))
        ttk.Button(row1, text="打开数据文件夹", command=self.open_data_folder).pack(side="left", padx=(10, 0))
        ttk.Button(row1, text="打开输出文件夹", command=self.open_output_folder).pack(side="left", padx=(10, 0))
        ttk.Button(row1, text="打开最新结果", command=self.open_latest_result).pack(side="left", padx=(10, 0))
        row2 = ttk.Frame(action_card, style="Toolbar.TFrame")
        row2.pack(fill="x", pady=(10, 0))
        ttk.Button(row2, text="导入申报表到历史库", command=self.import_current_declare).pack(side="left")
        ttk.Button(row2, text="打开历史库", command=self.open_history_folder).pack(side="left", padx=(10, 0))
        ttk.Button(row2, text="清除本页路径", style="Danger.TButton", command=self.clear_main_paths).pack(side="left", padx=(10, 0))
        ttk.Button(row2, text="退出程序", command=self.destroy).pack(side="right")

        config_card = ttk.Frame(config_page, style="Card.TFrame", padding=16)
        config_card.pack(fill="x", pady=(0, 12))
        ttk.Label(config_card, text="规则与高开票", style="Section.TLabel").grid(row=0, column=0, sticky="w", columnspan=3, pady=(0, 8))
        self._add_file_row(config_card, 1, "业务规则配置总表", self.config_var, lambda: self.choose_file("config"))
        self._add_file_row(config_card, 2, "高开票对账数据", self.high_invoice_var, lambda: self.choose_file("high_invoice"))

        config_buttons = ttk.Frame(config_page, style="Card.TFrame", padding=16)
        config_buttons.pack(fill="x", pady=(0, 12))
        ttk.Label(config_buttons, text="模板维护", style="Section.TLabel").pack(anchor="w", pady=(0, 10))
        config_row = ttk.Frame(config_buttons, style="Toolbar.TFrame")
        config_row.pack(fill="x")
        ttk.Button(config_row, text="导出默认配置表", command=self.export_default_config).pack(side="left")
        ttk.Button(config_row, text="打开模板文件夹", command=self.open_template_folder).pack(side="left", padx=(10, 0))
        ttk.Button(config_row, text="清除配置表路径", style="Danger.TButton", command=self.clear_config_path).pack(side="left", padx=(10, 0))
        ttk.Button(config_row, text="清除高开票路径", style="Danger.TButton", command=self.clear_high_invoice_path).pack(side="left", padx=(10, 0))

        help_text = (
            "业务规则配置总表：长期规则，包含特殊金额系数、银行1笔付款公司、医保折扣规则、基础参数。\n"
            "高开票对账数据：每次对账可能变化，单独上传，不并入长期配置。\n"
            "未选择配置表时，工具会使用模板文件夹里的默认配置；未选择高开票数据时，跳过高开票差额对账。"
        )
        help_card = ttk.Frame(config_page, style="Card.TFrame", padding=16)
        help_card.pack(fill="x", pady=(0, 12))
        ttk.Label(help_card, text="规则提示", style="Section.TLabel").pack(anchor="w", pady=(0, 8))
        ttk.Label(help_card, text=help_text, style="Hint.TLabel", justify="left").pack(anchor="w")

        log_frame = ttk.Frame(main_page, style="Card.TFrame", padding=16)
        log_frame.pack(fill="both", expand=True)
        ttk.Label(log_frame, text="运行日志", style="Section.TLabel").pack(anchor="w", pady=(0, 8))
        log_body = ttk.Frame(log_frame, style="Card.TFrame")
        log_body.pack(fill="both", expand=True)
        self.log_text = tk.Text(
            log_body,
            wrap="word",
            bg="#0F172A",
            fg="#E5E7EB",
            insertbackground="#E5E7EB",
            font=("Consolas", 10),
            relief="flat",
            padx=12,
            pady=12,
        )
        self.log_text.pack(side="left", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(log_body, orient="vertical", command=self.log_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=scrollbar.set)

    def _add_file_row(self, parent, row: int, label: str, var: tk.StringVar, command: Callable[[], None]) -> None:
        ttk.Label(parent, text=label, style="Card.TLabel").grid(row=row, column=0, sticky="w", padx=(0, 10), pady=6)
        entry = ttk.Entry(parent, textvariable=var)
        entry.grid(row=row, column=1, sticky="ew", pady=6)
        ttk.Button(parent, text="选择", command=command).grid(row=row, column=2, sticky="e", padx=(10, 0), pady=6)
        parent.columnconfigure(1, weight=1)

    def refresh_vars_from_selection(self) -> None:
        """把当前选择同步显示到界面输入框。"""

        self.drug_var.set(str(self.selection.drug_file or ""))
        self.bank_var.set(str(self.selection.bank_file or ""))
        self.declare_var.set(str(self.selection.declare_file or ""))
        self.output_var.set(str(self.selection.output_folder or ""))
        self.config_var.set(str(self.selection.config_file or ""))
        self.high_invoice_var.set(str(self.selection.high_invoice_file or ""))

    def sync_selection_from_vars(self) -> None:
        """从界面输入框同步路径，方便用户直接粘贴路径后执行。"""

        self.selection.drug_file = path_from_text(self.drug_var.get())
        self.selection.bank_file = path_from_text(self.bank_var.get())
        self.selection.declare_file = path_from_text(self.declare_var.get())
        self.selection.output_folder = path_from_text(self.output_var.get())
        self.selection.config_file = path_from_text(self.config_var.get())
        self.selection.high_invoice_file = path_from_text(self.high_invoice_var.get())

    def log(self, message: str) -> None:
        time_text = _dt.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert("end", f"[{time_text}] {message}\n")
        self.log_text.see("end")
        self.update_idletasks()

    def choose_file(self, kind: str) -> None:
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
        )
        if not file_path:
            return
        path = Path(file_path)
        if kind == "drug":
            self.selection.drug_file = path
            self.drug_var.set(str(path))
        elif kind == "bank":
            self.selection.bank_file = path
            self.bank_var.set(str(path))
        elif kind == "declare":
            self.selection.declare_file = path
            self.declare_var.set(str(path))
        elif kind == "config":
            self.selection.config_file = path
            self.config_var.set(str(path))
            self.log(f"已选择业务规则配置总表：{path.name}")
        elif kind == "high_invoice":
            self.selection.high_invoice_file = path
            self.high_invoice_var.set(str(path))
            self.log(f"已选择高开票对账数据：{path.name}")
        save_settings(self.selection)

    def choose_output_folder(self) -> None:
        folder = filedialog.askdirectory(title="选择结果输出文件夹")
        if not folder:
            return
        self.selection.output_folder = Path(folder)
        self.output_var.set(folder)
        self.log(f"结果将输出到：{folder}")
        save_settings(self.selection)

    def reload_default_folder(self) -> None:
        base = get_base_dir()
        folder = base / DEFAULT_DATA_FOLDER_NAME
        self.selection.data_folder = folder
        self.selection.output_folder = self.selection.output_folder or base
        self.output_var.set(str(self.selection.output_folder))
        self.load_files_from_folder(folder)
        save_settings(self.selection)

    def load_files_from_folder(self, folder: Path) -> None:
        """从一个文件夹加载默认命名的三个文件。"""

        self.selection.drug_file = folder / DEFAULT_DRUG_FILE
        self.selection.bank_file = folder / DEFAULT_BANK_FILE
        self.selection.declare_file = folder / DEFAULT_DECLARE_FILE

        self.drug_var.set(str(self.selection.drug_file))
        self.bank_var.set(str(self.selection.bank_file))
        self.declare_var.set(str(self.selection.declare_file))

        missing = [p.name for p in [self.selection.drug_file, self.selection.bank_file, self.selection.declare_file] if not p.exists()]
        if missing:
            self.log("请手动选择文件。默认位置未找到：" + "、".join(missing))
        else:
            self.log("已自动加载同级“对账数据表准备”中的三个 Excel，可以点击一键执行对账。")

    def clear_main_paths(self) -> None:
        """清空首页的药店、银行、申报表、输出文件夹路径。"""

        clear_main_file_settings(self.selection)
        self.refresh_vars_from_selection()
        self.log("已清空首页四个地址；规则配置和高开票路径未清空。")

    def clear_config_path(self) -> None:
        self.selection.config_file = None
        self.config_var.set("")
        save_settings(self.selection)
        self.log("已清除业务规则配置表路径，后续将使用默认配置。")

    def clear_high_invoice_path(self) -> None:
        self.selection.high_invoice_file = None
        self.high_invoice_var.set("")
        save_settings(self.selection)
        self.log("已清除高开票对账数据路径，后续将跳过高开票差额对账。")

    def open_template_folder(self) -> None:
        folder = get_base_dir() / TEMPLATE_FOLDER_NAME
        folder.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(folder)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"无法打开模板文件夹：{exc}")

    def export_default_config(self) -> None:
        """把默认配置总表另存一份给用户修改。"""

        source = get_default_config_path()
        if not source.exists():
            messagebox.showwarning(APP_TITLE, "模板文件夹中未找到默认配置表。")
            return
        target = filedialog.asksaveasfilename(
            title="导出默认配置表",
            defaultextension=".xlsx",
            initialfile=DEFAULT_CONFIG_FILE,
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not target:
            return
        try:
            shutil.copy2(source, target)
            self.log(f"已导出默认配置表：{target}")
            messagebox.showinfo(APP_TITLE, "默认配置表已导出，可以在 Excel 中修改后回到配置页上传。")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"导出配置表失败：{exc}")

    def execute(self) -> None:
        try:
            self.sync_selection_from_vars()
            save_settings(self.selection)
            self.log("========== 开始新一轮对账 ==========")
            output_path = run_reconcile(self.selection, self.log)
            self.last_output_path = output_path
            messagebox.showinfo(APP_TITLE, f"对账完成！\n\n结果文件：\n{output_path}")
        except UserVisibleError as exc:
            self.log(f"错误：{exc}")
            messagebox.showerror(APP_TITLE, str(exc))
        except PermissionError as exc:
            text = f"文件读写权限不足，请关闭已打开的结果表或检查文件夹权限。\n\n{exc}"
            self.log(text)
            messagebox.showerror(APP_TITLE, text)
        except Exception as exc:
            detail = traceback.format_exc()
            self.log("程序发生未预期错误：")
            self.log(detail)
            messagebox.showerror(APP_TITLE, f"程序发生未预期错误：{exc}\n\n详情已写入日志。")

    def open_output_folder(self) -> None:
        folder = self.selection.output_folder or get_base_dir()
        try:
            os.startfile(folder)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"无法打开文件夹：{exc}")

    def open_data_folder(self) -> None:
        folder = self.selection.data_folder or (get_base_dir() / DEFAULT_DATA_FOLDER_NAME)
        folder.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(folder)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"无法打开数据文件夹：{exc}")

    def open_latest_result(self) -> None:
        path = self.last_output_path
        if not path or not path.exists():
            output_folder = self.selection.output_folder or get_base_dir()
            candidates = sorted(output_folder.glob(f"{OUTPUT_PREFIX}_*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
            path = candidates[0] if candidates else None
        if not path or not path.exists():
            messagebox.showwarning(APP_TITLE, "还没有找到已生成的对账结果。")
            return
        try:
            os.startfile(path)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"无法打开最新结果：{exc}")

    def import_current_declare(self) -> None:
        try:
            if not self.selection.declare_file or not self.selection.declare_file.exists():
                messagebox.showwarning(APP_TITLE, "请先选择申报表文件。")
                return
            ym, target = import_declare_to_history(self.selection.declare_file)
            _, period = load_declare_medical_values(self.selection.declare_file)
            unit = period.get("unit") or "未知单位"
            self.log(f"已导入 {ym} {unit} 申报表到历史库：{target}")
            messagebox.showinfo(APP_TITLE, f"已导入 {ym} {unit} 申报表到历史库。")
        except Exception as exc:
            self.log(f"导入历史库失败：{exc}")
            messagebox.showerror(APP_TITLE, f"导入历史库失败：{exc}")

    def open_history_folder(self) -> None:
        folder = get_history_dir()
        folder.mkdir(parents=True, exist_ok=True)
        try:
            os.startfile(folder)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"无法打开历史库：{exc}")


def main() -> None:
    if tk is None:
        raise RuntimeError("当前环境不支持桌面界面，请使用 streamlit_app.py 启动网页版。")
    app = ReconcileApp()
    app.mainloop()


if __name__ == "__main__":
    main()
